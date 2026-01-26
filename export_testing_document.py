#!/usr/bin/env python
"""
Testing Document Generator
Exports all test cases to professional Excel format with detailed test steps
Template: Test Case ID | Function | Test Steps | Input Data | Expected Result
"""

import subprocess
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def run_all_tests():
    """Run all tests and return output"""
    print("Collecting all test cases...\n")
    
    result = subprocess.run(
        ['python', '-m', 'pytest', 'tests/', '--collect-only', '-q'],
        capture_output=True,
        text=True
    )
    
    return result.stdout + result.stderr


def parse_test_collection(output):
    """Parse pytest collection output to extract test information"""
    tests = []
    lines = output.split('\n')
    current_module = None
    current_class = None
    test_id_counter = 1
    
    for line in lines:
        # Extract module
        if '<Module ' in line:
            module_match = line.strip().split('<Module ')
            if len(module_match) > 1:
                current_module = module_match[1].rstrip('>')
        
        # Extract class
        elif '<Class ' in line:
            class_match = line.strip().split('<Class ')
            if len(class_match) > 1:
                current_class = class_match[1].rstrip('>')
        
        # Extract function
        elif '<Function test_' in line:
            func_match = line.strip().split('<Function ')
            if len(func_match) > 1:
                test_name = func_match[1].rstrip('>')
                
                # Determine test type and create ID
                if 'test_user_model' in current_module:
                    test_id = f"TM-UM-{test_id_counter:02d}"
                elif 'test_book_model' in current_module:
                    test_id = f"TM-BK-{test_id_counter:02d}"
                elif 'test_borrow_model' in current_module:
                    test_id = f"TM-BR-{test_id_counter:02d}"
                elif 'test_fine_model' in current_module:
                    test_id = f"TM-FN-{test_id_counter:02d}"
                elif 'test_' in current_module and 'controller' in current_module.lower():
                    test_id = f"TC-{test_id_counter:03d}"
                elif 'integration' in current_module:
                    test_id = f"TI-{test_id_counter:03d}"
                else:
                    test_id = f"TS-{test_id_counter:03d}"
                
                tests.append({
                    'id': test_id,
                    'module': current_module,
                    'class': current_class,
                    'name': test_name,
                    'function': test_name.replace('test_', '').replace('_', ' ').title()
                })
                
                test_id_counter += 1
    
    return tests


def get_test_details(test_name, test_class):
    """Get detailed test information: steps, input data, expected result"""
    
    test_details = {
        # USER MODEL TESTS
        'test_create_user_basic': {
            'steps': '1. Create User instance with name, email, password\n2. Verify name field initialized\n3. Verify email field initialized\n4. Verify password is hashed\n5. Verify role set to member\n6. Check is_active = True',
            'input': 'Name: John Doe\nEmail: john@example.com\nPassword: pass123\nRole: member',
            'result': 'User object created with all fields initialized'
        },
        'test_create_admin_user': {
            'steps': '1. Create User with admin role\n2. Set name and email\n3. Set password\n4. Verify role = admin\n5. Check is_active = True',
            'input': 'Name: Admin User\nEmail: admin@example.com\nPassword: admin123\nRole: admin',
            'result': 'Admin user created successfully'
        },
        'test_create_librarian_user': {
            'steps': '1. Create User with librarian role\n2. Set credentials\n3. Verify role = librarian\n4. Check is_active = True',
            'input': 'Name: Librarian\nEmail: lib@example.com\nPassword: lib123\nRole: librarian',
            'result': 'Librarian user created successfully'
        },
        'test_user_default_role': {
            'steps': '1. Create user without specifying role\n2. Check default role assigned\n3. Verify role = member\n4. Confirm explicit role required for admin/librarian',
            'input': 'Name: Test User\nEmail: test@example.com\nPassword: test123',
            'result': 'Default role = member assigned'
        },
        'test_save_user_to_database': {
            'steps': '1. Create user instance\n2. Hash password using pbkdf2_hmac\n3. Call user.save(db)\n4. Insert document to database\n5. Verify user_id assigned\n6. Check all fields persisted',
            'input': 'User object with complete data\nDatabase connection',
            'result': 'User saved to database with unique user_id'
        },
        'test_update_user_in_database': {
            'steps': '1. Create and save user\n2. Modify name and email fields\n3. Call save() again\n4. Update existing document\n5. Retrieve and verify changes',
            'input': 'Existing user\nNew name: Updated Name\nNew email: updated@example.com',
            'result': 'User record updated in database'
        },
        'test_delete_user_soft_delete': {
            'steps': '1. Create and save user\n2. Call user.delete()\n3. Set is_active = False\n4. Save to database\n5. Verify soft delete (not removed)',
            'input': 'Existing active user',
            'result': 'User marked as inactive, record still exists in database'
        },
        'test_get_user_by_id': {
            'steps': '1. Create and save user with user_id\n2. Call User.get_by_id(db, user_id)\n3. Query database by ObjectId\n4. Create User object from data\n5. Verify all fields match',
            'input': 'User ID (ObjectId)',
            'result': 'User object retrieved with correct data'
        },
        'test_get_user_by_email': {
            'steps': '1. Create and save user with email\n2. Call User.get_by_email(db, email)\n3. Query database by email\n4. Filter for is_active = True\n5. Verify user data matches',
            'input': 'User email',
            'result': 'Active user retrieved by email'
        },
        'test_get_nonexistent_user': {
            'steps': '1. Call get_by_id with invalid ID\n2. Verify None returned\n3. Call get_by_email with non-existent email\n4. Verify None returned\n5. Check no exception raised',
            'input': 'Invalid user_id or non-existent email',
            'result': 'None returned without error'
        },
        'test_get_all_users': {
            'steps': '1. Create multiple users\n2. Save all to database\n3. Call User.get_all(db)\n4. Query all active users\n5. Verify count matches',
            'input': 'Database with multiple active users',
            'result': 'All active users returned in list'
        },
        'test_get_users_by_role': {
            'steps': '1. Create users with different roles\n2. Save all to database\n3. Call get_by_role(db, role)\n4. Query by role field\n5. Verify only matching role returned',
            'input': 'Role: admin/librarian/member',
            'result': 'Only users with matching role returned'
        },
        'test_authenticate_valid_credentials': {
            'steps': '1. Create user with email and password\n2. Save to database\n3. Call User.authenticate(db, email, password)\n4. Hash provided password\n5. Compare with stored hash\n6. Verify authentication succeeds',
            'input': 'Email: user@example.com\nPassword: correct123',
            'result': 'User object returned, authentication successful'
        },
        'test_authenticate_invalid_password': {
            'steps': '1. Create user in database\n2. Call authenticate with wrong password\n3. Hash wrong password\n4. Compare with stored hash\n5. Verify authentication fails\n6. Return None',
            'input': 'Email: user@example.com\nPassword: wrongpass',
            'result': 'None returned, authentication failed'
        },
        'test_authenticate_nonexistent_email': {
            'steps': '1. Call authenticate with non-existent email\n2. Query database\n3. Verify no user found\n4. Return None\n5. Check no exception raised',
            'input': 'Email: nonexistent@example.com\nPassword: anypass',
            'result': 'None returned, no user found'
        },
        'test_user_with_empty_name': {
            'steps': '1. Try to create user with empty name\n2. Check validation error\n3. Verify exception raised\n4. Ensure user not created\n5. Check error message clear',
            'input': 'Name: "" (empty)\nEmail: test@example.com\nPassword: test123',
            'result': 'Validation error raised, user not created'
        },
        'test_user_with_empty_email': {
            'steps': '1. Try to create user with empty email\n2. Check validation error\n3. Verify exception raised\n4. Ensure user not created',
            'input': 'Name: Test User\nEmail: "" (empty)\nPassword: test123',
            'result': 'Validation error raised, user not created'
        },
        'test_user_email_case_sensitivity': {
            'steps': '1. Create user with email in uppercase\n2. Save to database\n3. Query with lowercase email\n4. Verify case-insensitive retrieval\n5. Create another user with different case',
            'input': 'Email: User@Example.COM\nQuery: user@example.com',
            'result': 'User retrieved regardless of case'
        },
        'test_get_user_borrow_records': {
            'steps': '1. Create user with borrows\n2. Create multiple borrow records\n3. Link to same user\n4. Call get_user_borrow_records\n5. Query all borrow records\n6. Filter by user_id',
            'input': 'User ID with multiple borrows',
            'result': 'All borrow records for user returned'
        },
        'test_get_user_active_borrows': {
            'steps': '1. Create user with returned and unreturned borrows\n2. Mark some as returned\n3. Call get_user_active_borrows\n4. Filter where return_date is None\n5. Verify only active returned',
            'input': 'User with mixed borrow status',
            'result': 'Only unreturned borrow records returned'
        },

        # BOOK MODEL TESTS
        'test_create_book_basic': {
            'steps': '1. Create Book with title, author, publisher, year, category\n2. Verify title field\n3. Verify author field\n4. Check default quantity = 1\n5. Check default available = 1\n6. Verify is_active = True',
            'input': 'Title: Python Programming\nAuthor: Guido van Rossum\nPublisher: Tech Books\nYear: 2023\nCategory: Programming',
            'result': 'Book object created with all fields initialized'
        },
        'test_create_book_with_quantity': {
            'steps': '1. Create book with quantity parameter\n2. Set quantity = 5\n3. Set available = 3\n4. Verify quantity field\n5. Verify available field\n6. Check available <= quantity',
            'input': 'Title: Test Book\nQuantity: 5\nAvailable: 3',
            'result': 'Book created with specified quantity and availability'
        },
        'test_book_is_active_default': {
            'steps': '1. Create book instance\n2. Check is_active property\n3. Verify default = True\n4. Ensure book searchable by default',
            'input': 'Book object',
            'result': 'is_active = True by default'
        },
        'test_save_book_to_database': {
            'steps': '1. Create book instance\n2. Call book.save(db)\n3. Insert into database\n4. Verify book_id assigned\n5. Check document persisted\n6. Verify all fields saved',
            'input': 'Book object, database connection',
            'result': 'Book saved with unique book_id'
        },
        'test_update_book_in_database': {
            'steps': '1. Create and save book\n2. Modify title and quantity\n3. Call save() again\n4. Update existing document\n5. Verify changes persisted',
            'input': 'Existing book, new title, new quantity',
            'result': 'Book record updated in database'
        },
        'test_delete_book_soft_delete': {
            'steps': '1. Create and save book\n2. Call book.delete()\n3. Set is_active = False\n4. Save to database\n5. Verify soft delete',
            'input': 'Active book',
            'result': 'Book marked inactive, not removed from database'
        },
        'test_get_book_by_id': {
            'steps': '1. Create and save book\n2. Get book_id\n3. Call Book.get_by_id(db, book_id)\n4. Query database\n5. Create Book object\n6. Verify all fields match',
            'input': 'Book ID (ObjectId)',
            'result': 'Book object retrieved with correct data'
        },
        'test_get_nonexistent_book': {
            'steps': '1. Create empty database\n2. Call Book.get_by_id with invalid ID\n3. Verify None returned\n4. Check no exception raised',
            'input': 'Invalid book_id',
            'result': 'None returned'
        },
        'test_get_all_books': {
            'steps': '1. Create multiple books\n2. Save all to database\n3. Call Book.get_all(db)\n4. Query all active books\n5. Verify count matches',
            'input': 'Database with multiple books',
            'result': 'All active books returned'
        },
        'test_search_books_by_title': {
            'steps': '1. Create books with different titles\n2. Save to database\n3. Call search_by_title(db, keyword)\n4. Query title field\n5. Support partial match\n6. Verify matching books returned',
            'input': 'Search keyword: "Python"',
            'result': 'All books with matching title returned'
        },
        'test_search_books_by_author': {
            'steps': '1. Create books with different authors\n2. Save to database\n3. Call search_by_author(db, author)\n4. Query author field\n5. Support partial match\n6. Verify matching books returned',
            'input': 'Author name: "Guido"',
            'result': 'All books by matching author returned'
        },
        'test_search_books_by_category': {
            'steps': '1. Create books in different categories\n2. Save to database\n3. Call get_by_category(db, category)\n4. Query by category field\n5. Verify all matching returned',
            'input': 'Category: "Programming"',
            'result': 'All books in category returned'
        },
        'test_check_book_available': {
            'steps': '1. Create book with quantity and available\n2. Check available > 0\n3. Verify is_available() method\n4. Return True if available\n5. Return False if available = 0',
            'input': 'Book with available = 3',
            'result': 'is_available() returns True'
        },
        'test_reduce_availability': {
            'steps': '1. Create book with available = 5\n2. Call reduce_availability()\n3. Decrease by 1\n4. Verify available = 4\n5. Save to database',
            'input': 'Book with available = 5',
            'result': 'Available decreased to 4'
        },
        'test_increase_availability': {
            'steps': '1. Create book with available = 3\n2. Call increase_availability()\n3. Increase by 1\n4. Verify available = 4\n5. Check not exceed quantity',
            'input': 'Book with available = 3, quantity = 5',
            'result': 'Available increased to 4'
        },
        'test_book_with_zero_year': {
            'steps': '1. Create book with year = 0\n2. Check validation\n3. Verify exception or warning\n4. Set reasonable year\n5. Verify creation succeeds',
            'input': 'Year: 0',
            'result': 'Validation handled appropriately'
        },
        'test_book_with_negative_quantity': {
            'steps': '1. Try to create book with quantity < 0\n2. Check validation error\n3. Verify exception raised\n4. Ensure book not created',
            'input': 'Quantity: -5',
            'result': 'Validation error raised'
        },
        'test_book_with_empty_title': {
            'steps': '1. Try to create book with empty title\n2. Check validation error\n3. Verify exception raised\n4. Ensure book not created',
            'input': 'Title: "" (empty)',
            'result': 'Validation error raised'
        },
        'test_count_total_books': {
            'steps': '1. Create multiple books\n2. Save all to database\n3. Call count_total(db)\n4. Count active books\n5. Verify count correct',
            'input': 'Database with 10 books',
            'result': 'Total count = 10'
        },
        'test_get_books_by_category': {
            'steps': '1. Create books in different categories\n2. Save to database\n3. Call get_by_category(db, category)\n4. Query by category field\n5. Verify all matching returned',
            'input': 'Category: "Fiction"',
            'result': 'All fiction books returned'
        },

        # BORROW MODEL TESTS
        'test_create_borrow_record': {
            'steps': '1. Create borrow with user_id and book_id\n2. Set borrow_date\n3. Calculate due_date (14 days)\n4. Create BorrowRecord instance\n5. Verify all fields initialized',
            'input': 'User ID, Book ID',
            'result': 'BorrowRecord created with borrow/due dates'
        },
        'test_borrow_record_default_duration': {
            'steps': '1. Create borrow record\n2. Check default borrow duration\n3. Verify 14 days duration\n4. Due date = borrow_date + 14 days\n5. Verify dates calculated',
            'input': 'BorrowRecord without custom duration',
            'result': 'Default 14-day duration applied'
        },
        'test_save_borrow_record': {
            'steps': '1. Create borrow instance\n2. Call borrow.save(db)\n3. Insert into database\n4. Verify borrow_id assigned\n5. Check document persisted',
            'input': 'BorrowRecord object, database',
            'result': 'Borrow saved with unique borrow_id'
        },
        'test_update_borrow_record': {
            'steps': '1. Create and save borrow\n2. Set return_date\n3. Call save() again\n4. Update existing document\n5. Verify changes persisted',
            'input': 'Existing borrow, return_date = today',
            'result': 'Borrow record updated with return date'
        },
        'test_get_borrow_record_by_id': {
            'steps': '1. Create and save borrow\n2. Get borrow_id\n3. Call BorrowRecord.get_by_id(db, borrow_id)\n4. Query database\n5. Create BorrowRecord object\n6. Verify all fields match',
            'input': 'Borrow ID (ObjectId)',
            'result': 'BorrowRecord retrieved with correct data'
        },
        'test_get_user_borrow_records': {
            'steps': '1. Create multiple borrow records\n2. Link to same user_id\n3. Call get_user_records(db, user_id)\n4. Query all user borrows\n5. Verify all returned',
            'input': 'User ID with multiple borrows',
            'result': 'All borrow records for user returned'
        },
        'test_get_book_borrow_records': {
            'steps': '1. Create multiple borrow records\n2. Link to same book_id\n3. Call get_book_records(db, book_id)\n4. Query all book borrows\n5. Verify all returned',
            'input': 'Book ID with multiple borrows',
            'result': 'All borrow records for book returned'
        },
        'test_is_active_borrow_record': {
            'steps': '1. Create borrow record\n2. Check is_active() method\n3. Return True if return_date = None\n4. Return False if return_date set\n5. Test both states',
            'input': 'BorrowRecord with/without return_date',
            'result': 'is_active() returns correct boolean'
        },
        'test_is_overdue_borrow_record': {
            'steps': '1. Create borrow with past due_date\n2. Call is_overdue()\n3. Compare due_date with today\n4. Return True if overdue\n5. Test both scenarios',
            'input': 'BorrowRecord with past/future due_date',
            'result': 'is_overdue() returns correct boolean'
        },
        'test_mark_borrow_as_returned': {
            'steps': '1. Create active borrow\n2. Call mark_as_returned()\n3. Set return_date = today\n4. Update status\n5. Save to database',
            'input': 'Active borrow record',
            'result': 'Borrow marked as returned with return_date set'
        },
        'test_calculate_days_borrowed': {
            'steps': '1. Create borrow with past borrow_date\n2. Call get_days_borrowed()\n3. Calculate days between dates\n4. Return integer\n5. Verify calculation accurate',
            'input': 'BorrowRecord from past',
            'result': 'Days borrowed calculated correctly'
        },
        'test_calculate_days_until_due': {
            'steps': '1. Create borrow with future due_date\n2. Call get_days_until_due()\n3. Calculate days remaining\n4. Return positive if not overdue\n5. Test both scenarios',
            'input': 'BorrowRecord with future/past due_date',
            'result': 'Days until due calculated correctly'
        },
        'test_borrow_record_with_custom_due_date': {
            'steps': '1. Create borrow with custom due_date\n2. Override default 14 days\n3. Set specific due date\n4. Save to database\n5. Verify custom date preserved',
            'input': 'Custom due_date parameter',
            'result': 'Custom due date applied instead of default'
        },
        'test_borrow_record_with_empty_user_id': {
            'steps': '1. Try create borrow with empty user_id\n2. Check validation error\n3. Verify exception raised\n4. Ensure record not created',
            'input': 'User ID: "" (empty)',
            'result': 'Validation error raised'
        },
        'test_borrow_record_with_empty_book_id': {
            'steps': '1. Try create borrow with empty book_id\n2. Check validation error\n3. Verify exception raised\n4. Ensure record not created',
            'input': 'Book ID: "" (empty)',
            'result': 'Validation error raised'
        },
        'test_borrow_record_return_after_borrow': {
            'steps': '1. Create borrow record\n2. Verify return_date = None\n3. Mark as returned\n4. Set return_date\n5. Verify borrow marked complete',
            'input': 'BorrowRecord',
            'result': 'Return date set, borrow marked complete'
        },
        'test_get_active_borrows': {
            'steps': '1. Create multiple borrows\n2. Mark some as returned\n3. Call get_active(db, user_id)\n4. Filter where return_date = None\n5. Verify only active returned',
            'input': 'Mixed active/returned borrows',
            'result': 'Only unreturned borrows returned'
        },
        'test_get_overdue_borrows': {
            'steps': '1. Create borrows with past due dates\n2. Create on-time borrows\n3. Call get_overdue(db)\n4. Compare due_date with today\n5. Verify only overdue returned',
            'input': 'Borrows with past/future due dates',
            'result': 'Only overdue borrows returned'
        },
        'test_get_returned_borrows': {
            'steps': '1. Create multiple borrows\n2. Mark some as returned\n3. Call get_returned(db)\n4. Filter where return_date != None\n5. Verify only returned returned',
            'input': 'Mixed active/returned borrows',
            'result': 'Only returned borrows returned'
        },

        # FINE MODEL TESTS
        'test_create_fine_basic': {
            'steps': '1. Create fine with borrow_record_id\n2. Input fine amount\n3. Set user_id\n4. Create Fine instance\n5. Verify record_id field\n6. Check default status = unpaid',
            'input': 'Record ID, Amount: 50.00, User ID',
            'result': 'Fine object created with unpaid status'
        },
        'test_create_fine_default_status': {
            'steps': '1. Create fine instance\n2. Check default status\n3. Verify status = unpaid\n4. Ensure can mark as paid',
            'input': 'Fine object',
            'result': 'Default status = unpaid'
        },
        'test_create_fine_with_amount': {
            'steps': '1. Create fine with amount parameter\n2. Set amount = 150.50\n3. Verify amount field\n4. Check currency handling',
            'input': 'Amount: 150.50',
            'result': 'Fine created with specified amount'
        },
        'test_save_fine_to_database': {
            'steps': '1. Create fine instance\n2. Call fine.save(db)\n3. Insert into database\n4. Verify fine_id assigned\n5. Check document persisted',
            'input': 'Fine object, database',
            'result': 'Fine saved with unique fine_id'
        },
        'test_update_fine_in_database': {
            'steps': '1. Create and save fine\n2. Update amount field\n3. Call save() again\n4. Update existing document\n5. Verify changes persisted',
            'input': 'Existing fine, new amount',
            'result': 'Fine record updated in database'
        },
        'test_mark_fine_as_paid': {
            'steps': '1. Create unpaid fine\n2. Call mark_as_paid()\n3. Set status = paid\n4. Set payment_date = today\n5. Save to database',
            'input': 'Unpaid fine',
            'result': 'Fine marked as paid with payment_date'
        },
        'test_get_fine_by_id': {
            'steps': '1. Create and save fine\n2. Get fine_id\n3. Call Fine.get_by_id(db, fine_id)\n4. Query database\n5. Create Fine object\n6. Verify all fields match',
            'input': 'Fine ID (ObjectId)',
            'result': 'Fine object retrieved with correct data'
        },
        'test_get_all_fines': {
            'steps': '1. Create multiple fines\n2. Save all to database\n3. Call Fine.get_all(db)\n4. Query all fines\n5. Verify all returned',
            'input': 'Database with multiple fines',
            'result': 'All fines returned'
        },
        'test_get_nonexistent_fine': {
            'steps': '1. Call Fine.get_by_id with invalid ID\n2. Verify None returned\n3. Check no exception raised',
            'input': 'Invalid fine_id',
            'result': 'None returned'
        },
        'test_fine_pending_status': {
            'steps': '1. Create fine\n2. Check status = unpaid\n3. Verify is_paid() = False\n4. Verify is_pending() = True',
            'input': 'Newly created fine',
            'result': 'is_pending() returns True'
        },
        'test_fine_paid_status': {
            'steps': '1. Create fine\n2. Mark as paid\n3. Check status = paid\n4. Verify is_paid() = True',
            'input': 'Paid fine',
            'result': 'is_paid() returns True'
        },
        'test_get_pending_fines': {
            'steps': '1. Create fines with paid and unpaid\n2. Call get_pending(db)\n3. Filter where status = unpaid\n4. Verify only pending returned',
            'input': 'Mixed paid/unpaid fines',
            'result': 'Only unpaid fines returned'
        },
        'test_fine_amount_calculation': {
            'steps': '1. Create fine with calculated amount\n2. Verify amount > 0\n3. Check decimal precision\n4. Verify currency format',
            'input': 'Overdue days: 5, rate: 10000/day',
            'result': 'Fine amount = 50000 calculated correctly'
        },
        'test_fine_zero_amount': {
            'steps': '1. Create fine with amount = 0\n2. Check validation\n3. May allow or error',
            'input': 'Amount: 0',
            'result': 'Handled according to business logic'
        },
        'test_fine_high_amount': {
            'steps': '1. Create fine with large amount\n2. Test precision\n3. Verify decimal handling',
            'input': 'Amount: 999999.99',
            'result': 'Large amount handled correctly'
        },
        'test_payment_date_on_pay': {
            'steps': '1. Create unpaid fine\n2. Mark as paid\n3. Verify payment_date = today\n4. Check timestamp\n5. Verify persisted',
            'input': 'Unpaid fine',
            'result': 'Payment date set to current date'
        },
        'test_payment_status_update': {
            'steps': '1. Create fine\n2. Check initial status = unpaid\n3. Mark as paid\n4. Verify status = paid\n5. Update in database',
            'input': 'Fine object',
            'result': 'Status updated from unpaid to paid'
        },
        'test_fine_with_empty_record_id': {
            'steps': '1. Try create fine with empty record_id\n2. Check validation\n3. Verify exception raised',
            'input': 'Record ID: "" (empty)',
            'result': 'Validation error raised'
        },
        'test_fine_negative_amount': {
            'steps': '1. Try create fine with amount < 0\n2. Check validation\n3. Verify exception raised',
            'input': 'Amount: -100',
            'result': 'Validation error raised'
        },
        'test_get_total_fines_amount': {
            'steps': '1. Create multiple fines\n2. Sum all amounts\n3. Call get_total_amount(db)\n4. Calculate total\n5. Verify calculation',
            'input': 'Multiple fines with amounts: 50, 100, 75',
            'result': 'Total = 225 calculated correctly'
        },
        'test_get_fines_by_status': {
            'steps': '1. Create fines with different status\n2. Call get_by_status(db, status)\n3. Filter by status field\n4. Verify all matching returned',
            'input': 'Status: paid/unpaid',
            'result': 'Only fines with matching status returned'
        },
    }
    
    # Return default if not found
    if test_name in test_details:
        return test_details[test_name]
    else:
        return {
            'steps': 'Execute test steps',
            'input': 'Test inputs',
            'result': 'Expected result'
        }


def create_excel_file(tests):
    """Create Excel workbook with testing document template"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Define styles
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True, size=12, color="000000")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Test Case ID", "Function", "Test Steps", "Input Data", "Expected Result"]
    ws.append(headers)
    
    # Format header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    
    # Add test data
    for idx, test in enumerate(tests, 1):
        test_id = test['id']
        function_name = test['function']
        details = get_test_details(test['name'], test['class'])
        
        ws.append([
            test_id,
            function_name,
            details['steps'],
            details['input'],
            details['result']
        ])
        
        # Format data row
        for col_num in range(1, 6):
            cell = ws.cell(row=idx + 1, column=col_num)
            cell.border = border
            cell.alignment = left_alignment
    
    # Set row height for headers
    ws.row_dimensions[1].height = 20
    
    # Auto-fit row height for data rows (with minimum)
    for row_num in range(2, len(tests) + 2):
        ws.row_dimensions[row_num].height = None  # Auto
    
    return wb


def main():
    # Collect tests
    output = run_all_tests()
    tests = parse_test_collection(output)
    
    if not tests:
        print("ERROR: No tests found!")
        return
    
    print(f"Found {len(tests)} test cases")
    print(f"Creating Testing Document Excel file...\n")
    
    # Create Excel file
    wb = create_excel_file(tests)
    wb.save("Testing_Document.xlsx")
    
    print("✓ Testing Document exported to: Testing_Document.xlsx")
    print()
    print("=" * 50)
    print("Testing Document Summary")
    print("=" * 50)
    print(f"Total Test Cases:     {len(tests)}")
    print(f"Format:               Template-based (TCID|Function|Steps|Input|Result)")
    print(f"File Size:            {len(tests)} rows + header")
    print(f"Generated:            {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 50)
    print("\nColumns:")
    print("  A: Test Case ID      (Unique identifier)")
    print("  B: Function          (What is being tested)")
    print("  C: Test Steps        (Detailed step-by-step)")
    print("  D: Input Data        (Test inputs)")
    print("  E: Expected Result   (What should happen)")
    print("=" * 50)


if __name__ == '__main__':
    main()
