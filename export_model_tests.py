#!/usr/bin/env python
"""
Model Tests Export to Excel
Exports all model unit tests to professional Excel format
"""

import subprocess
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def run_model_tests():
    """Run all model tests and return output"""
    print("Starting model test export to Excel...\n")
    print("Running pytest tests/unit/test_*_model.py...\n")
    
    result = subprocess.run(
        ['python', '-m', 'pytest', 
         'tests/unit/test_user_model.py',
         'tests/unit/test_book_model.py', 
         'tests/unit/test_borrow_model.py',
         'tests/unit/test_fine_model.py',
         '-v', '--tb=no', '--no-header'],
        capture_output=True,
        text=True
    )
    
    return result.stdout + result.stderr


def parse_test_output(output):
    """Parse pytest output to extract test results"""
    tests = []
    lines = output.split('\n')
    
    for line in lines:
        # Match test result lines: test_method PASSED/FAILED/SKIPPED
        # Pattern: tests/unit/test_*.py::ClassName::test_method PASSED
        if 'test_' in line and ('PASSED' in line or 'FAILED' in line or 'SKIPPED' in line):
            # Extract test name from format: tests/unit/test_user_model.py::ClassName::test_name PASSED
            parts = line.split('::')
            if len(parts) >= 3:
                test_name = parts[-1].split()[0]  # Get test name before status
                status = 'PASSED' if 'PASSED' in line else ('FAILED' if 'FAILED' in line else 'SKIPPED')
                tests.append({
                    'name': test_name,
                    'status': status
                })
    
    return tests


def get_test_description(test_method):
    """Get detailed test description with step-by-step breakdown"""
    descriptions = {
        # User Model Tests
        'test_create_user_basic': '1. Input user details (name, email, password)\n2. Set role to member\n3. Create User instance\n4. Verify name field\n5. Verify email field\n6. Verify password is hashed\n7. Check is_active = True',
        'test_create_admin_user': '1. Create user with admin role\n2. Set name and email\n3. Set password\n4. Set role = admin\n5. Verify role property\n6. Check is_active = True',
        'test_create_librarian_user': '1. Create user with librarian role\n2. Set credentials\n3. Set role = librarian\n4. Verify role property\n5. Check is_active = True',
        'test_user_default_role': '1. Create user without specifying role\n2. Check default role is member\n3. Verify role property\n4. Ensure explicit role required for admin/librarian',
        'test_save_user_to_database': '1. Create user instance\n2. Hash password using pbkdf2_hmac\n3. Call user.save(db)\n4. Insert document into database\n5. Verify user_id assigned\n6. Check document persisted\n7. Verify all fields saved',
        'test_update_user_in_database': '1. Create and save user\n2. Modify user fields (name, email)\n3. Call save() again\n4. Update existing document\n5. Verify changes persisted\n6. Retrieve and confirm update',
        'test_delete_user_soft_delete': '1. Create and save user\n2. Call user.delete()\n3. Set is_active = False\n4. Save to database\n5. Verify soft delete (not removed)\n6. Check record still exists\n7. Verify is_active = False',
        'test_get_user_by_id': '1. Create and save user\n2. Get user_id\n3. Call User.get_by_id(db, user_id)\n4. Query database by ObjectId\n5. Create User object from data\n6. Verify all fields match\n7. Check password hash preserved',
        'test_get_user_by_email': '1. Create and save user with email\n2. Call User.get_by_email(db, email)\n3. Query database by email\n4. Filter for is_active = True\n5. Verify only active users returned\n6. Check user data matches\n7. Verify email case-insensitive',
        'test_get_nonexistent_user': '1. Create empty database\n2. Call User.get_by_id with invalid ID\n3. Verify None returned\n4. Try get_by_email with non-existent email\n5. Verify None returned\n6. Check no exception raised',
        'test_get_all_users': '1. Create multiple users\n2. Save all to database\n3. Call User.get_all(db)\n4. Query all active users\n5. Filter is_active = True\n6. Verify all users returned\n7. Check count matches',
        'test_get_users_by_role': '1. Create users with different roles\n2. Save all to database\n3. Call get_by_role(db, role)\n4. Query by role field\n5. Verify only matching role returned\n6. Check count correct\n7. Test all role types',
        'test_authenticate_valid_credentials': '1. Create user with email and password\n2. Save to database\n3. Call User.authenticate(db, email, password)\n4. Hash provided password\n5. Compare with stored hash\n6. Verify authentication succeeds\n7. Return user object',
        'test_authenticate_invalid_password': '1. Create user in database\n2. Call authenticate with wrong password\n3. Hash wrong password\n4. Compare with stored hash\n5. Verify authentication fails\n6. Return None\n7. Check no exception raised',
        'test_authenticate_nonexistent_email': '1. Create empty database\n2. Call authenticate with non-existent email\n3. Query database\n4. Verify no user found\n5. Return None\n6. Check no exception raised',
        'test_user_with_empty_name': '1. Try to create user with empty name\n2. Check validation error\n3. Verify exception raised\n4. Ensure user not created\n5. Check error message clear',
        'test_user_with_empty_email': '1. Try to create user with empty email\n2. Check validation error\n3. Verify exception raised\n4. Ensure user not created\n5. Check error message clear',
        'test_user_email_case_sensitivity': '1. Create user with email in uppercase\n2. Save to database\n3. Query with lowercase email\n4. Verify case-insensitive retrieval\n5. Create another user with different case\n6. Verify duplicate prevention\n7. Check email stored as-is',
        'test_get_user_borrow_records': '1. Create user with borrows\n2. Create multiple borrow records\n3. Link to same user\n4. Call get_user_borrow_records\n5. Query all borrow records\n6. Filter by user_id\n7. Verify all returned',
        'test_get_user_active_borrows': '1. Create user with returned and unreturned borrows\n2. Mark some as returned\n3. Call get_user_active_borrows\n4. Filter where return_date is None\n5. Verify only active returned\n6. Check count matches\n7. Verify returned excluded',

        # Book Model Tests
        'test_create_book_basic': '1. Input book details (title, author, publisher, year, category)\n2. Create Book instance\n3. Verify title field\n4. Verify author field\n5. Check default quantity = 1\n6. Check default available = 1\n7. Verify is_active = True',
        'test_create_book_with_quantity': '1. Create book with quantity parameter\n2. Set quantity = 5\n3. Set available = 3\n4. Verify quantity field\n5. Verify available field\n6. Check relationship (available <= quantity)\n7. Ensure valid state',
        'test_book_is_active_default': '1. Create book instance\n2. Check is_active property\n3. Verify default = True\n4. Ensure book searchable by default\n5. Check status field initialized',
        'test_save_book_to_database': '1. Create book instance\n2. Call book.save(db)\n3. Insert into database\n4. Verify book_id assigned\n5. Check document persisted\n6. Verify all fields saved\n7. Confirm searchable after save',
        'test_update_book_in_database': '1. Create and save book\n2. Modify fields (title, quantity)\n3. Call save() again\n4. Update existing document\n5. Verify changes persisted\n6. Retrieve and confirm\n7. Check version updated',
        'test_delete_book_soft_delete': '1. Create and save book\n2. Call book.delete()\n3. Set is_active = False\n4. Save to database\n5. Verify soft delete\n6. Check record still exists\n7. Verify not in active search',
        'test_get_book_by_id': '1. Create and save book\n2. Get book_id\n3. Call Book.get_by_id(db, book_id)\n4. Query database by ObjectId\n5. Create Book object from data\n6. Verify all fields match\n7. Check title and author preserved',
        'test_get_nonexistent_book': '1. Create empty database\n2. Call Book.get_by_id with invalid ID\n3. Verify None returned\n4. Try search by title\n5. Verify None returned\n6. Check no exception raised',
        'test_get_all_books': '1. Create multiple books\n2. Save all to database\n3. Call Book.get_all(db)\n4. Query all active books\n5. Filter is_active = True\n6. Verify all returned\n7. Check count matches',
        'test_search_books_by_title': '1. Create books with different titles\n2. Save to database\n3. Call search_by_title(db, keyword)\n4. Query title field\n5. Support partial match\n6. Verify matching books returned\n7. Filter is_active = True',
        'test_search_books_by_author': '1. Create books with different authors\n2. Save to database\n3. Call search_by_author(db, author)\n4. Query author field\n5. Support partial match\n6. Verify matching books returned\n7. Filter active books',
        'test_search_books_by_category': '1. Create books in different categories\n2. Save to database\n3. Call get_by_category(db, category)\n4. Query by category field\n5. Verify all matching category returned\n6. Check count correct\n7. Test multiple categories',
        'test_check_book_available': '1. Create book with quantity and available\n2. Check available > 0\n3. Verify is_available() method\n4. Return True if available\n5. Return False if available = 0\n6. Test edge cases\n7. Verify availability logic',
        'test_reduce_availability': '1. Create book with available = 5\n2. Call reduce_availability()\n3. Decrease by 1\n4. Verify available = 4\n5. Save to database\n6. Verify persisted\n7. Test cannot go below 0',
        'test_increase_availability': '1. Create book with available = 3\n2. Call increase_availability()\n3. Increase by 1\n4. Verify available = 4\n5. Check not exceed quantity\n6. Save to database\n7. Verify persisted',
        'test_book_with_zero_year': '1. Create book with year = 0\n2. Check validation\n3. Verify exception or warning\n4. Set reasonable year\n5. Verify creation succeeds\n6. Check error handling',
        'test_book_with_negative_quantity': '1. Try to create book with quantity < 0\n2. Check validation error\n3. Verify exception raised\n4. Ensure book not created\n5. Check error message',
        'test_book_with_empty_title': '1. Try to create book with empty title\n2. Check validation error\n3. Verify exception raised\n4. Ensure book not created\n5. Check error clear',
        'test_count_total_books': '1. Create multiple books\n2. Save all to database\n3. Call count_total(db)\n4. Count active books\n5. Filter is_active = True\n6. Verify count correct\n7. Test pagination',
        'test_get_books_by_category': '1. Create books in different categories\n2. Save to database\n3. Call get_by_category(db, category)\n4. Query by category field\n5. Verify all matching returned\n6. Check count matches\n7. Test multiple categories',

        # Borrow Model Tests
        'test_create_borrow_record': '1. Create borrow with user_id and book_id\n2. Set borrow_date\n3. Calculate due_date (14 days)\n4. Create BorrowRecord instance\n5. Verify user_id field\n6. Verify book_id field\n7. Check return_date = None initially',
        'test_borrow_record_default_duration': '1. Create borrow record\n2. Check default borrow duration\n3. Verify 14 days duration\n4. Due date = borrow_date + 14 days\n5. Verify dates calculated\n6. Check timezone handling',
        'test_save_borrow_record': '1. Create borrow instance\n2. Call borrow.save(db)\n3. Insert into database\n4. Verify borrow_id assigned\n5. Check document persisted\n6. Verify all fields saved\n7. Confirm dates preserved',
        'test_update_borrow_record': '1. Create and save borrow\n2. Set return_date\n3. Call save() again\n4. Update existing document\n5. Verify changes persisted\n6. Retrieve and confirm\n7. Check status updated',
        'test_get_borrow_record_by_id': '1. Create and save borrow\n2. Get borrow_id\n3. Call BorrowRecord.get_by_id(db, borrow_id)\n4. Query database\n5. Create BorrowRecord object\n6. Verify all fields match\n7. Check dates preserved',
        'test_get_user_borrow_records': '1. Create multiple borrow records\n2. Link to same user_id\n3. Call get_user_records(db, user_id)\n4. Query all user borrows\n5. Verify all returned\n6. Check count matches\n7. Sort by date',
        'test_get_book_borrow_records': '1. Create multiple borrow records\n2. Link to same book_id\n3. Call get_book_records(db, book_id)\n4. Query all book borrows\n5. Verify all returned\n6. Check count matches\n7. Include returned and active',
        'test_is_active_borrow_record': '1. Create borrow record\n2. Check is_active() method\n3. Return True if return_date = None\n4. Return False if return_date set\n5. Test both states\n6. Verify logic correct',
        'test_is_overdue_borrow_record': '1. Create borrow with past due_date\n2. Call is_overdue()\n3. Compare due_date with today\n4. Return True if overdue\n5. Return False if on time\n6. Create returned borrow\n7. Test both scenarios',
        'test_mark_borrow_as_returned': '1. Create active borrow\n2. Call mark_as_returned()\n3. Set return_date = today\n4. Update status\n5. Save to database\n6. Verify persisted\n7. Check is_active = False',
        'test_calculate_days_borrowed': '1. Create borrow with past borrow_date\n2. Call get_days_borrowed()\n3. Calculate days between dates\n4. Return integer\n5. Test current borrows (0 days)\n6. Test old borrows\n7. Verify calculation accurate',
        'test_calculate_days_until_due': '1. Create borrow with future due_date\n2. Call get_days_until_due()\n3. Calculate days remaining\n4. Return positive if not overdue\n5. Return negative if overdue\n6. Test both scenarios\n7. Verify calculation',
        'test_borrow_record_with_custom_due_date': '1. Create borrow with custom due_date\n2. Override default 14 days\n3. Set specific due date\n4. Save to database\n5. Retrieve and verify\n6. Check custom date preserved\n7. Calculate days correctly',
        'test_borrow_record_with_empty_user_id': '1. Try create borrow with empty user_id\n2. Check validation error\n3. Verify exception raised\n4. Ensure record not created\n5. Check error message',
        'test_borrow_record_with_empty_book_id': '1. Try create borrow with empty book_id\n2. Check validation error\n3. Verify exception raised\n4. Ensure record not created\n5. Check error message',
        'test_borrow_record_return_after_borrow': '1. Create borrow record\n2. Verify return_date = None\n3. Mark as returned\n4. Set return_date\n5. Verify borrow marked complete\n6. Calculate days borrowed\n7. Update availability',
        'test_get_active_borrows': '1. Create multiple borrows\n2. Mark some as returned\n3. Call get_active(db, user_id)\n4. Filter where return_date = None\n5. Verify only active returned\n6. Check returned excluded\n7. Sort by due_date',
        'test_get_overdue_borrows': '1. Create borrows with past due dates\n2. Create on-time borrows\n3. Call get_overdue(db)\n4. Compare due_date with today\n5. Filter overdue\n6. Verify only overdue returned\n7. Check on-time excluded',
        'test_get_returned_borrows': '1. Create multiple borrows\n2. Mark some as returned\n3. Call get_returned(db)\n4. Filter where return_date != None\n5. Verify only returned\n6. Check count matches\n7. Sort by return_date',

        # Fine Model Tests
        'test_create_fine_basic': '1. Create fine with borrow_record_id\n2. Input fine amount\n3. Set user_id\n4. Create Fine instance\n5. Verify record_id field\n6. Verify amount field\n7. Check default status = unpaid',
        'test_create_fine_default_status': '1. Create fine instance\n2. Check default status\n3. Verify status = unpaid\n4. Ensure can mark as paid\n5. Verify status property',
        'test_create_fine_with_amount': '1. Create fine with amount parameter\n2. Set amount = 150.50\n3. Verify amount field\n4. Check currency handling\n5. Verify decimal precision',
        'test_save_fine_to_database': '1. Create fine instance\n2. Call fine.save(db)\n3. Insert into database\n4. Verify fine_id assigned\n5. Check document persisted\n6. Verify all fields saved\n7. Confirm searchable',
        'test_update_fine_in_database': '1. Create and save fine\n2. Update amount field\n3. Call save() again\n4. Update existing document\n5. Verify changes persisted\n6. Retrieve and confirm\n7. Check update_date set',
        'test_mark_fine_as_paid': '1. Create unpaid fine\n2. Call mark_as_paid()\n3. Set status = paid\n4. Set payment_date = today\n5. Save to database\n6. Verify persisted\n7. Check status = paid',
        'test_get_fine_by_id': '1. Create and save fine\n2. Get fine_id\n3. Call Fine.get_by_id(db, fine_id)\n4. Query database\n5. Create Fine object\n6. Verify all fields match\n7. Check amount preserved',
        'test_get_all_fines': '1. Create multiple fines\n2. Save all to database\n3. Call Fine.get_all(db)\n4. Query all fines\n5. Verify all returned\n6. Check count matches\n7. Filter active fines',
        'test_get_nonexistent_fine': '1. Create empty database\n2. Call Fine.get_by_id with invalid ID\n3. Verify None returned\n4. Check no exception raised',
        'test_fine_pending_status': '1. Create fine\n2. Check status = unpaid\n3. Verify is_paid() = False\n4. Verify is_pending() = True\n5. Test status property',
        'test_fine_paid_status': '1. Create fine\n2. Mark as paid\n3. Check status = paid\n4. Verify is_paid() = True\n5. Verify is_pending() = False',
        'test_get_pending_fines': '1. Create fines with paid and unpaid\n2. Call get_pending(db)\n3. Filter where status = unpaid\n4. Verify only pending returned\n5. Check paid excluded\n6. Sort by date',
        'test_fine_amount_calculation': '1. Create fine with calculated amount\n2. Verify amount > 0\n3. Check decimal precision\n4. Verify currency format\n5. Test rounding logic',
        'test_fine_zero_amount': '1. Create fine with amount = 0\n2. Check validation\n3. May allow or error\n4. Verify handling\n5. Check business logic',
        'test_fine_high_amount': '1. Create fine with large amount\n2. Test precision\n3. Verify decimal handling\n4. Check currency format\n5. Test limits',
        'test_payment_date_on_pay': '1. Create unpaid fine\n2. Mark as paid\n3. Set payment_date\n4. Verify payment_date = today\n5. Check timestamp\n6. Verify persisted',
        'test_payment_status_update': '1. Create fine\n2. Check initial status = unpaid\n3. Mark as paid\n4. Verify status = paid\n5. Update in database\n6. Retrieve and confirm\n7. Check immutable',
        'test_fine_with_empty_record_id': '1. Try create fine with empty record_id\n2. Check validation\n3. Verify exception raised\n4. Ensure fine not created\n5. Check error message',
        'test_fine_negative_amount': '1. Try create fine with amount < 0\n2. Check validation\n3. Verify exception raised\n4. Ensure fine not created\n5. Check error clear',
        'test_get_total_fines_amount': '1. Create multiple fines\n2. Sum all amounts\n3. Call get_total_amount(db)\n4. Calculate total\n5. Filter by user or date\n6. Verify calculation\n7. Test aggregation',
        'test_get_fines_by_status': '1. Create fines with different status\n2. Call get_by_status(db, status)\n3. Filter by status field\n4. Verify all matching returned\n5. Test paid and unpaid\n6. Check count matches',
    }
    
    return descriptions.get(test_method, f"Execute {test_method.replace('test_', '').replace('_', ' ').title()}")


def create_excel_file(tests):
    """Create Excel workbook with test results"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Tests"
    
    # Define styles
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    passed_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    failed_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(bold=True, size=12, color="000000")
    status_font_passed = Font(bold=True, color="FFFFFF")
    status_font_failed = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Test ID", "Test Method", "Status", "Test Steps"]
    ws.append(headers)
    
    # Format header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 60
    
    # Add test data
    for idx, test in enumerate(tests, 1):
        test_id = f"TM-{idx:03d}"
        test_method = test['name']
        status = test['status']
        description = get_test_description(test_method)
        
        ws.append([test_id, test_method, status, description])
        
        # Format data row
        for col_num in range(1, 5):
            cell = ws.cell(row=idx + 1, column=col_num)
            cell.border = border
            
            if col_num == 3:  # Status column
                cell.alignment = center_alignment
                if status == "PASSED":
                    cell.fill = passed_fill
                    cell.font = status_font_passed
                elif status == "FAILED":
                    cell.fill = failed_fill
                    cell.font = status_font_failed
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    
    total_tests = len(tests)
    passed_tests = sum(1 for t in tests if t['status'] == 'PASSED')
    failed_tests = sum(1 for t in tests if t['status'] == 'FAILED')
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    
    summary_data = [
        ["Model Test Summary", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total Tests", total_tests],
        ["Passed", passed_tests],
        ["Failed", failed_tests],
        ["Pass Rate", f"{pass_rate:.1f}%"],
        ["", ""],
        ["Test Coverage", ""],
        ["User Model", sum(1 for t in tests if 'test_' in t['name'] and any(x in t['name'] for x in ['create_user', 'save_user', 'get_user', 'authenticate', 'delete_user', 'user_borrow', 'user_with']))],
        ["Book Model", sum(1 for t in tests if 'test_' in t['name'] and any(x in t['name'] for x in ['create_book', 'save_book', 'get_book', 'search_books', 'book_available', 'book_with', 'count_total', 'get_books_by_category']))],
        ["Borrow Model", sum(1 for t in tests if 'test_' in t['name'] and any(x in t['name'] for x in ['create_borrow', 'borrow_record', 'save_borrow', 'get_borrow', 'is_active_borrow', 'is_overdue_borrow', 'mark_borrow', 'calculate_days', 'get_active_borrows', 'get_overdue_borrows', 'get_returned_borrows']))],
        ["Fine Model", sum(1 for t in tests if 'test_' in t['name'] and any(x in t['name'] for x in ['create_fine', 'save_fine', 'get_fine', 'fine_', 'payment', 'get_total_fines', 'get_fines_by_status']))],
    ]
    
    for row in summary_data:
        summary_ws.append(row)
    
    # Format summary sheet
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 15
    
    return wb


def main():
    # Run tests
    output = run_model_tests()
    
    # Parse results
    tests = parse_test_output(output)
    
    if not tests:
        print("ERROR: No tests found!")
        return
    
    print(f"Found {len(tests)} model tests")
    
    # Group by status
    passed = sum(1 for t in tests if t['status'] == 'PASSED')
    failed = sum(1 for t in tests if t['status'] == 'FAILED')
    
    print(f"Creating Excel file...")
    
    # Create Excel file
    wb = create_excel_file(tests)
    wb.save("Model_Test_Results.xlsx")
    
    print("✓ Test results exported to: Model_Test_Results.xlsx")
    print()
    print("=" * 40)
    print("Model Test Summary")
    print("=" * 40)
    print(f"Total Tests:  {len(tests)}")
    print(f"Passed:       {passed}")
    print(f"Failed:       {failed}")
    print(f"Pass Rate:    {(passed/len(tests)*100):.1f}%")
    print("=" * 40)


if __name__ == '__main__':
    main()
