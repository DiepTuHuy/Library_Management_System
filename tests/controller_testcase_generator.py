"""
Controller Test Case Generator
Automatically extracts and generates comprehensive test case documentation
from controller test files and exports to Excel format
"""

import os
import re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ControllerTestCaseExtractor:
    """Extract test cases from controller test files"""
    
    def __init__(self, test_dir):
        self.test_dir = test_dir
        self.test_cases = []
    
    def extract_test_methods(self, file_path):
        """Extract test method names and docstrings from test file"""
        tests = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Find all test methods and their docstrings
            pattern = r'def (test_\w+)\(self\):\s*(?:"""(.*?)"""|\'\'\'(.*?)\'\'\')?'
            matches = re.finditer(pattern, content, re.DOTALL)
            
            for match in matches:
                test_name = match.group(1)
                docstring = match.group(2) or match.group(3) or ""
                tests.append({
                    'name': test_name,
                    'docstring': docstring.strip()
                })
        except Exception as e:
            print(f"Error reading {file_path}: {e}")
        
        return tests
    
    def extract_from_directory(self):
        """Extract test cases from all controller test files"""
        controllers = {
            'test_user_controller.py': {'prefix': 'UC', 'controller': 'User Controller'},
            'test_book_controller.py': {'prefix': 'BC', 'controller': 'Book Controller'},
            'test_borrow_controller.py': {'prefix': 'BRC', 'controller': 'Borrow Controller'},
            'test_fine_controller.py': {'prefix': 'FC', 'controller': 'Fine Controller'},
        }
        
        test_counter = {}
        
        for filename, info in controllers.items():
            filepath = os.path.join(self.test_dir, 'unit', filename)
            if os.path.exists(filepath):
                tests = self.extract_test_methods(filepath)
                prefix = info['prefix']
                test_counter[prefix] = 1
                
                for test in tests:
                    test_id = f"{prefix}-{test_counter[prefix]:03d}"
                    test_counter[prefix] += 1
                    
                    self.test_cases.append({
                        'Test ID': test_id,
                        'Controller': info['controller'],
                        'Test Method': test['name'],
                        'Docstring': test['docstring'],
                        'Description': self.generate_description(test['name']),
                        'Preconditions': self.generate_preconditions(test['name']),
                        'Test Steps': self.generate_test_steps(test['name']),
                        'Expected Result': self.generate_expected_result(test['name']),
                    })
        
        return self.test_cases
    
    def generate_description(self, test_name):
        """Generate test description from test method name"""
        # Convert test_method_name to "Test Method Name"
        name = test_name.replace('test_', '').replace('_', ' ')
        return name.capitalize()
    
    def generate_preconditions(self, test_name):
        """Generate preconditions based on test name"""
        if 'invalid' in test_name or 'error' in test_name:
            return "Invalid data or error condition exists"
        elif 'nonexistent' in test_name or 'not_exist' in test_name:
            return "Entity does not exist in database"
        elif 'duplicate' in test_name:
            return "Duplicate entry exists"
        elif 'empty' in test_name:
            return "Empty data or fields"
        else:
            return "Database is ready, valid data exists"
    
    def generate_test_steps(self, test_name):
        """Generate generic test steps"""
        steps = [
            "1. Execute the test method with prepared data",
            "2. Verify method execution completed",
            "3. Check response/result for correctness",
            "4. Validate data state after execution"
        ]
        return "\n".join(steps)
    
    def generate_expected_result(self, test_name):
        """Generate expected result based on test name"""
        if 'success' in test_name or 'valid' in test_name:
            return "Operation completed successfully with correct output"
        elif 'fail' in test_name or 'error' in test_name or 'invalid' in test_name:
            return "Operation failed with appropriate error/exception"
        elif 'nonexistent' in test_name or 'not_exist' in test_name:
            return "Returns None or empty result for non-existent entity"
        elif 'empty' in test_name:
            return "Empty results returned or validation error raised"
        else:
            return "Test assertion passes with expected behavior"


class TestCaseExcelGenerator:
    """Generate Excel workbook from test cases"""
    
    def __init__(self, test_cases, output_path):
        self.test_cases = test_cases
        self.output_path = output_path
    
    def create_workbook(self):
        """Create Excel workbook with test case data"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Controller Test Cases'
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Create headers
        headers = [
            'Test ID', 
            'Controller', 
            'Test Method', 
            'Description', 
            'Preconditions', 
            'Test Steps', 
            'Expected Result',
            'Status'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Add test cases
        for row_idx, test_case in enumerate(self.test_cases, 2):
            ws.cell(row=row_idx, column=1, value=test_case['Test ID']).border = border
            ws.cell(row=row_idx, column=2, value=test_case['Controller']).border = border
            ws.cell(row=row_idx, column=3, value=test_case['Test Method']).border = border
            ws.cell(row=row_idx, column=4, value=test_case['Description']).border = border
            ws.cell(row=row_idx, column=5, value=test_case['Preconditions']).border = border
            ws.cell(row=row_idx, column=6, value=test_case['Test Steps']).border = border
            ws.cell(row=row_idx, column=7, value=test_case['Expected Result']).border = border
            ws.cell(row=row_idx, column=8, value='').border = border
            
            # Apply alignment and wrapping to all cells
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 12
        
        # Set row heights
        ws.row_dimensions[1].height = 30
        for row_idx in range(2, len(self.test_cases) + 2):
            ws.row_dimensions[row_idx].height = 80
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        return wb
    
    def save(self):
        """Save workbook to file"""
        wb = self.create_workbook()
        wb.save(self.output_path)
        return self.output_path


class TestCaseReport:
    """Generate summary report of test cases"""
    
    def __init__(self, test_cases):
        self.test_cases = test_cases
    
    def generate_summary(self):
        """Generate text summary of test cases"""
        summary = []
        summary.append("=" * 80)
        summary.append("CONTROLLER TEST CASE SUMMARY")
        summary.append("=" * 80)
        summary.append("")
        
        # Group by controller
        by_controller = {}
        for test_case in self.test_cases:
            controller = test_case['Controller']
            if controller not in by_controller:
                by_controller[controller] = []
            by_controller[controller].append(test_case)
        
        # Print by controller
        total_tests = 0
        for controller in sorted(by_controller.keys()):
            tests = by_controller[controller]
            summary.append(f"\n{controller}:")
            summary.append("-" * 80)
            summary.append(f"Total Tests: {len(tests)}")
            summary.append("")
            
            for test in tests:
                summary.append(f"  [{test['Test ID']}] {test['Test Method']}")
                summary.append(f"      Description: {test['Description']}")
            
            total_tests += len(tests)
        
        summary.append("")
        summary.append("=" * 80)
        summary.append(f"TOTAL TEST CASES: {total_tests}")
        summary.append("=" * 80)
        summary.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        summary.append("")
        
        return "\n".join(summary)
    
    def save_summary(self, output_path):
        """Save summary to text file"""
        summary = self.generate_summary()
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(summary)
        return output_path


class ControllerTestCaseGenerator:
    """Main generator class"""
    
    def __init__(self, project_root):
        self.project_root = project_root
        self.test_dir = os.path.join(project_root, 'tests')
        self.output_dir = os.path.join(project_root, 'Testing_docs')
        
        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate(self):
        """Generate test case documentation"""
        print("\n" + "=" * 80)
        print("CONTROLLER TEST CASE GENERATOR")
        print("=" * 80)
        
        # Extract test cases
        print("\nExtracting test cases from controller test files...")
        extractor = ControllerTestCaseExtractor(self.test_dir)
        test_cases = extractor.extract_from_directory()
        
        if not test_cases:
            print("❌ No test cases found!")
            return
        
        print(f"✓ Extracted {len(test_cases)} test cases")
        
        # Generate Excel file
        print("\nGenerating Excel documentation...")
        excel_path = os.path.join(self.output_dir, 'controller_testcase_reference.xlsx')
        excel_generator = TestCaseExcelGenerator(test_cases, excel_path)
        excel_generator.save()
        print(f"✓ Excel file saved: {excel_path}")
        
        # Generate summary report
        print("\nGenerating summary report...")
        report = TestCaseReport(test_cases)
        summary_path = os.path.join(self.test_dir, 'controller_testcase_summary.txt')
        report.save_summary(summary_path)
        print(f"✓ Summary report saved: {summary_path}")
        
        # Print summary to console
        print("\n" + report.generate_summary())
        
        print("=" * 80)
        print("✓ Test case documentation generation complete!")
        print("=" * 80 + "\n")


def main():
    """Main entry point"""
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    generator = ControllerTestCaseGenerator(project_root)
    generator.generate()


if __name__ == '__main__':
    main()
