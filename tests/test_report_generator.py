"""
Test Report Generator - Template Format
Generates Excel test documentation in template format from pytest tests
"""

import os
import re
import time
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class TestReportGenerator:
    """Generate test reports in Excel template format"""
    
    def __init__(self, project_root):
        self.project_root = project_root
        self.tests_dir = os.path.join(project_root, 'tests')
        self.output_dir = os.path.join(project_root, 'Testing_docs')
        self.all_test_cases = []
        self.controller_test_cases = []
        self.model_test_cases = []
        self.tc_id_counter = 1
        os.makedirs(self.output_dir, exist_ok=True)
    
    def parse_test_file(self, file_path):
        """Parse test file to extract test cases in template format"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Extract test methods with their docstrings
            method_pattern = r'def (test_\w+)\(self[^)]*\):\s*"""([^"]*)"""'
            methods = re.findall(method_pattern, content)
            
            # Determine if this is controller or model test
            is_controller = 'controller' in file_path
            
            # Extract feature/module name
            feature_module = self._extract_feature_module(file_path)
            
            for method_name, method_doc in methods:
                test_case = {
                    'Test ID': f'TC-{self.tc_id_counter:03d}',
                    'Feature / Module': feature_module,
                    'Test Case Description': method_doc.strip() if method_doc else 'Test case',
                    'Preconditions': self._extract_preconditions(content, method_name),
                    'Test Steps': self._format_test_steps(method_doc),
                    'Expected Result': self._extract_expected_result(content, method_name),
                    'Actual Results': 'Pending',  # Default value for testing
                    'Status': 'Not Run'  # Default status
                }
                
                self.all_test_cases.append(test_case)
                
                # Categorize by controller or model
                if is_controller:
                    self.controller_test_cases.append(test_case)
                else:
                    self.model_test_cases.append(test_case)
                
                self.tc_id_counter += 1
        
        except Exception as e:
            print(f"Error parsing {file_path}: {e}")
    
    def _extract_function_name(self, file_path, method_name):
        """Extract function name from file and method"""
        filename = os.path.basename(file_path).replace('.py', '')
        # Extract meaningful name (e.g., test_book_model -> Book, test_user_controller -> User)
        if 'test_' in filename:
            parts = filename.replace('test_', '').replace('_model', '').replace('_controller', '')
            return parts.title()
        return filename
    
    def _extract_feature_module(self, file_path):
        """Extract feature/module name from file path"""
        filename = os.path.basename(file_path).replace('.py', '')
        
        # For controller files
        if 'controller' in filename:
            # Extract controller name (e.g., test_auth_controller -> auth_controller)
            if 'test_' in filename:
                module_name = filename.replace('test_', '')
            else:
                module_name = filename
            
            # Format as controllers.module_name (ClassName)
            parts = module_name.replace('_controller', '').split('_')
            class_name = ''.join([p.capitalize() for p in parts]) + 'Controller'
            return f'controllers.{module_name} ({class_name})'
        
        # For model files
        elif 'model' in filename:
            if 'test_' in filename:
                module_name = filename.replace('test_', '')
            else:
                module_name = filename
            
            parts = module_name.replace('_model', '').split('_')
            class_name = ''.join([p.capitalize() for p in parts])
            return f'models.{module_name} ({class_name})'
        
        return filename
    
    def _extract_preconditions(self, content, method_name):
        """Extract preconditions from test method"""
        pattern = rf'def {method_name}\(self[^)]*\):'
        match = re.search(pattern, content)
        
        if match:
            # Extract parameters from method signature
            params_pattern = rf'def {method_name}\(self([^)]*)\):'
            params_match = re.search(params_pattern, content)
            
            if params_match:
                params = params_match.group(1).strip()
                if params:
                    # Convert params to preconditions
                    param_list = [p.strip() for p in params.split(',')]
                    preconditions = []
                    for i, param in enumerate(param_list, 1):
                        if 'mock' in param or 'sample' in param:
                            preconditions.append(f'{i}. {param} fixture initialized')
                    
                    if preconditions:
                        return '\n'.join(preconditions)
        
        return 'Test fixtures initialized'
    
    def _format_test_steps(self, doc_string):
        """Format docstring as numbered test steps (1., 2., 3.,...)"""
        if not doc_string:
            # Default steps for generic tests
            return "1. Initialize test data\n2. Perform test action\n3. Verify expected result"
        
        # Clean up the docstring
        doc_string = doc_string.strip()
        
        # Check if already has numbered format
        if '1.' in doc_string or '2.' in doc_string:
            # Already formatted, just clean it up
            lines = doc_string.split('\n')
            steps = [line.strip() for line in lines if line.strip()]
            return '\n'.join(steps)
        
        # If docstring has multiple lines, convert to numbered steps
        if '\n' in doc_string:
            lines = [line.strip() for line in doc_string.split('\n') if line.strip()]
            numbered_steps = '\n'.join([f'{i+1}. {line}' for i, line in enumerate(lines)])
            return numbered_steps
        
        # Single line docstring - create 3 generic steps
        return f"1. {doc_string}\n2. Verify result\n3. Assert expected outcome"
    
    def _extract_input_data(self, content, method_name):
        """Extract input data from test method"""
        pattern = rf'def {method_name}\(.*?\n(.*?)(?=\n    def|\nclass|\Z)'
        match = re.search(pattern, content, re.DOTALL)
        
        input_data = []
        
        if match:
            method_body = match.group(1)
            
            # Extract variable assignments with string values
            var_pattern = r'(\w+)\s*=\s*["\']([^"\']*)["\']'
            vars_found = re.findall(var_pattern, method_body)
            for var_name, var_value in vars_found[:5]:  # Limit to first 5
                input_data.append(f'{var_name}: {var_value}')
            
            # Extract method parameters from fixture
            param_pattern = r'(sample_\w+|mock_\w+)'
            params = re.findall(param_pattern, method_body)
            for param in set(params):
                if param not in ' '.join(input_data):
                    input_data.insert(0, param)
        
        # If still empty, generate generic input
        if not input_data:
            input_data.append('Test data fixtures')
        
        return '\n'.join(input_data) if input_data else 'Standard test data'
    
    def _extract_expected_result(self, content, method_name):
        """Extract expected result from assertions in test method"""
        pattern = rf'def {method_name}\(.*?\n(.*?)(?=\n    def|\nclass|\Z)'
        match = re.search(pattern, content, re.DOTALL)
        
        expected_results = []
        
        if match:
            method_body = match.group(1)
            # Look for assertions
            assertions = re.findall(r'assert\s+(.+?)(?:\n|;|$)', method_body)
            for assertion in assertions[:3]:  # Limit to first 3 assertions
                # Simplify assertion for readability
                assertion = assertion.strip()
                if ' is True' in assertion:
                    expected_results.append('Should return True')
                elif ' is False' in assertion:
                    expected_results.append('Should return False')
                elif ' is not None' in assertion:
                    expected_results.append('Should not be None')
                elif ' is None' in assertion:
                    expected_results.append('Should be None')
                elif '>' in assertion or '<' in assertion:
                    expected_results.append(f'Assert: {assertion[:60]}')
                elif '==' in assertion:
                    expected_results.append(f'Assert: {assertion[:60]}')
                else:
                    expected_results.append(assertion[:70])
        
        # If no assertions found, create default
        if not expected_results:
            expected_results.append('Test passes without errors')
        
        return '\n'.join(expected_results) if expected_results else 'Test passes without errors'
    
    def scan_test_files(self):
        """Scan all test files"""
        test_files = []
        
        # Scan unit test directory
        unit_test_dir = os.path.join(self.tests_dir, 'unit')
        if os.path.exists(unit_test_dir):
            for file in sorted(os.listdir(unit_test_dir)):
                if file.startswith('test_') and file.endswith('.py'):
                    test_files.append(os.path.join(unit_test_dir, file))
        
        # Also check root test files
        test_controller_file = os.path.join(self.tests_dir, 'test_controller.py')
        if os.path.exists(test_controller_file):
            test_files.append(test_controller_file)
        
        # Parse all test files
        for test_file in test_files:
            print(f"  Parsing {os.path.basename(test_file)}...")
            self.parse_test_file(test_file)
    
    def create_excel_workbook(self, test_cases=None, columns=None):
        """Create Excel workbook with template format"""
        if test_cases is None:
            test_cases = self.all_test_cases
        if columns is None:
            columns = ['Test ID', 'Feature / Module', 'Test Case Description', 'Preconditions', 'Test Steps', 'Expected Result', 'Actual Results', 'Status']
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Cases'
        
        # Define header style - Yellow background
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True, size=11, color="000000")
        
        # Define border style
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Add header row
        for col_idx, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Add test case data
        for row_idx, test_case in enumerate(test_cases, 2):
            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=test_case.get(column, ''))
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Set column widths
        col_widths = {
            'Test ID': 12,
            'Feature / Module': 32,
            'Test Case Description': 35,
            'Preconditions': 30,
            'Test Steps': 35,
            'Expected Result': 35,
            'Actual Results': 35,
            'Status': 12,
            'Tester': 15,
            'Test Case ID': 12,
            'Function': 18,
            'Test Steps': 35,
            'Input Data': 25,
        }
        
        for col_idx, column in enumerate(columns, 1):
            width = col_widths.get(column, 20)
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, len(test_cases) + 2):
            ws.row_dimensions[row_idx].height = 60
        
        return wb
    
    def generate_excel_reports(self):
        """Generate all Excel reports"""
        print("\n" + "="*60)
        print("Generating Test Reports in Template Format")
        print("="*60)
        
        print("\nScanning test files...")
        self.scan_test_files()
        
        if not self.all_test_cases:
            print("No test cases found!")
            return
        
        print(f"\nFound {len(self.all_test_cases)} test cases")
        print(f"  - Controllers: {len(self.controller_test_cases)}")
        print(f"  - Models: {len(self.model_test_cases)}")
        
        # Create controller test case report with new template format
        if self.controller_test_cases:
            print("\nCreating controller_testcase.xlsx...")
            columns_controller = ['Test ID', 'Feature / Module', 'Test Case Description', 'Preconditions', 'Test Steps', 'Expected Result', 'Actual Results', 'Status']
            wb = self.create_excel_workbook(self.controller_test_cases, columns_controller)
            
            controller_path = os.path.join(self.output_dir, 'controller_testcase.xlsx')
            self._save_excel_file(wb, controller_path)
            print(f"✓ Created: {controller_path}")
        
        # Create testing document with simplified template format
        print("\nCreating testing_document.xlsx...")
        # Convert all test cases to simple format for testing_document
        test_cases_simple = self._convert_to_simple_format(self.all_test_cases)
        columns_basic = ['Test Case ID', 'Function', 'Test Steps', 'Input Data', 'Expected Result']
        wb = self.create_excel_workbook(test_cases_simple, columns_basic)
        
        testing_doc_path = os.path.join(self.output_dir, 'testing_document.xlsx')
        self._save_excel_file(wb, testing_doc_path)
        print(f"✓ Created: {testing_doc_path}")
        
        print(f"\n✓ All reports generated in: {self.output_dir}")
        print("="*60 + "\n")
    
    def _save_excel_file(self, wb, file_path):
        """Save Excel file with retry logic for file lock"""
        max_retries = 3
        for attempt in range(max_retries):
            try:
                wb.save(file_path)
                return
            except PermissionError:
                if attempt < max_retries - 1:
                    print(f"  File locked, retrying in 2 seconds...")
                    time.sleep(2)
                else:
                    print(f"✗ Could not save file after {max_retries} attempts")
                    raise
    
    def _convert_to_simple_format(self, test_cases):
        """Convert test cases to simple format for testing_document"""
        simple_cases = []
        tc_id = 1
        
        for test_case in test_cases:
            simple_case = {
                'Test Case ID': f'TC{tc_id:02d}',
                'Function': test_case.get('Test Case Description', test_case.get('Function', '')),
                'Test Steps': test_case.get('Test Steps', ''),
                'Input Data': test_case.get('Preconditions', test_case.get('Input Data', '')),
                'Expected Result': test_case.get('Expected Result', '')
            }
            simple_cases.append(simple_case)
            tc_id += 1
        
        return simple_cases


def main():
    """Main entry point"""
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    generator = TestReportGenerator(project_root)
    generator.generate_excel_reports()


if __name__ == '__main__':
    main()
