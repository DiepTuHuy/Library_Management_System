"""
Test Failure Report Generator
Generates Excel test failure report for failed tests
"""

import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class TestFailureReportGenerator:
    """Generate test failure report in Excel format"""
    
    def __init__(self, project_root):
        self.project_root = project_root
        self.output_dir = os.path.join(project_root, 'Testing_docs')
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Define failed tests with error information
        self.failed_tests = [
            {
                'Test ID': 'TC-001',
                'Test Name': 'test_librarian_manages_borrows_and_returns',
                'Module': 'test_system_workflows',
                'Error Type': 'AttributeError',
                'Error Message': "'BorrowController' object has no attribute 'return_book'",
                'Root Cause': 'Method name mismatch in controller implementation',
                'Fix': 'Use correct method name: get_all_active_borrows() instead of get_active_borrows()',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-002',
                'Test Name': 'test_admin_manages_users',
                'Module': 'test_system_workflows',
                'Error Type': 'AttributeError',
                'Error Message': "'BorrowRecord' is not defined",
                'Root Cause': 'Missing import statement in test file',
                'Fix': 'Add: from models.borrow_record import BorrowRecord',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-003',
                'Test Name': 'test_borrow_record_creation',
                'Module': 'test_borrow_controller',
                'Error Type': 'NameError',
                'Error Message': "name 'BorrowRecord' is not defined",
                'Root Cause': 'Missing import in test file',
                'Fix': 'Add: from models.borrow_record import BorrowRecord',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-004',
                'Test Name': 'test_overdue_check',
                'Module': 'test_borrow_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'BorrowRecord' object has no attribute 'return_date'",
                'Root Cause': 'Attribute name mismatch in model',
                'Fix': 'Use correct attribute: return_date (check model definition)',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-005',
                'Test Name': 'test_get_user_borrows',
                'Module': 'test_borrow_controller',
                'Error Type': 'AssertionError',
                'Error Message': "assert 'borrows' in result",
                'Root Cause': 'API response key mismatch - returns records not borrows',
                'Fix': 'Update test: use result[\"records\"] instead of result[\"borrows\"]',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-006',
                'Test Name': 'test_get_active_borrows',
                'Module': 'test_borrow_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'BorrowController' object has no attribute 'get_active_borrows'",
                'Root Cause': 'Method name mismatch',
                'Fix': 'Use: get_all_active_borrows() instead of get_active_borrows()',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-007',
                'Test Name': 'test_get_overdue_borrows',
                'Module': 'test_borrow_controller',
                'Error Type': 'AssertionError',
                'Error Message': "assert 'borrows' in result",
                'Root Cause': 'API response key mismatch',
                'Fix': 'Update test: check for \"records\" key instead of \"borrows\"',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-008',
                'Test Name': 'test_create_fine_success',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'FineController' object has no attribute 'create_fine'",
                'Root Cause': 'Method does not exist in FineController',
                'Fix': 'Implement create_fine() method or update test to use existing methods',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-009',
                'Test Name': 'test_create_fine_for_overdue_book',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'FineController' object has no attribute 'create_fine'",
                'Root Cause': 'Method does not exist in FineController',
                'Fix': 'Implement create_fine() method or update test',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-010',
                'Test Name': 'test_create_fine_invalid_record',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'FineController' object has no attribute 'create_fine'",
                'Root Cause': 'Method does not exist in FineController',
                'Fix': 'Implement create_fine() method',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-011',
                'Test Name': 'test_calculate_overdue_days',
                'Module': 'test_fine_controller',
                'Error Type': 'NameError',
                'Error Message': "name 'BorrowRecord' is not defined",
                'Root Cause': 'Missing import in test file',
                'Fix': 'Add: from models.borrow_record import BorrowRecord',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-012',
                'Test Name': 'test_get_unpaid_fines',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'FineController' object has no attribute 'get_unpaid_fines'",
                'Root Cause': 'Method name mismatch',
                'Fix': 'Use: get_all_pending_fines() instead of get_unpaid_fines()',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-013',
                'Test Name': 'test_fine_creation',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'Fine' object has no attribute 'is_paid'",
                'Root Cause': 'Fine model uses status=\"paid/pending\", not is_paid boolean',
                'Fix': 'Update test: use fine.status==\"paid\" instead of fine.is_paid',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-014',
                'Test Name': 'test_fine_status',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'Fine' object has no attribute 'is_paid'",
                'Root Cause': 'Fine model uses status field, not is_paid property',
                'Fix': 'Update test: check fine.status instead of fine.is_paid',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-015',
                'Test Name': 'test_pay_fine_success',
                'Module': 'test_fine_controller',
                'Error Type': 'TypeError',
                'Error Message': "FineController.pay_fine() got unexpected keyword argument 'payment_method'",
                'Root Cause': 'pay_fine() method does not accept payment_method parameter',
                'Fix': 'Remove payment_method parameter from test or implement in controller',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-016',
                'Test Name': 'test_pay_fine_partial_amount',
                'Module': 'test_fine_controller',
                'Error Type': 'TypeError',
                'Error Message': "FineController.pay_fine() got unexpected keyword argument 'payment_method'",
                'Root Cause': 'Method signature mismatch',
                'Fix': 'Update method call to match FineController.pay_fine() signature',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-017',
                'Test Name': 'test_pay_nonexistent_fine',
                'Module': 'test_fine_controller',
                'Error Type': 'TypeError',
                'Error Message': "FineController.pay_fine() got unexpected keyword argument 'payment_method'",
                'Root Cause': 'Method signature mismatch',
                'Fix': 'Update method call signature',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-018',
                'Test Name': 'test_pay_already_paid_fine',
                'Module': 'test_fine_controller',
                'Error Type': 'TypeError',
                'Error Message': "FineController.pay_fine() got unexpected keyword argument 'payment_method'",
                'Root Cause': 'Method signature mismatch',
                'Fix': 'Update method call signature',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-019',
                'Test Name': 'test_payment_creation',
                'Module': 'test_fine_controller',
                'Error Type': 'TypeError',
                'Error Message': "Payment.__init__() got unexpected keyword argument 'payment_method'",
                'Root Cause': 'Payment model does not accept payment_method parameter',
                'Fix': 'Check Payment model __init__ signature and update test accordingly',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-020',
                'Test Name': 'test_payment_save_and_retrieve',
                'Module': 'test_fine_controller',
                'Error Type': 'TypeError',
                'Error Message': "Payment.__init__() got unexpected keyword argument 'payment_method'",
                'Root Cause': 'Payment constructor signature mismatch',
                'Fix': 'Update test to use correct Payment constructor parameters',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-021',
                'Test Name': 'test_fine_is_unpaid',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'Fine' object has no attribute 'is_paid'",
                'Root Cause': 'Fine model does not have is_paid attribute',
                'Fix': 'Update test: use fine.status == \"pending\" instead',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-022',
                'Test Name': 'test_fine_marks_as_paid',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'Fine' object has no attribute 'is_paid'",
                'Root Cause': 'Fine model uses status field',
                'Fix': 'Update test: check retrieved.status == \"paid\"',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-023',
                'Test Name': 'test_total_pending_fines',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'FineController' object has no attribute 'get_total_pending_fines'",
                'Root Cause': 'Method name mismatch',
                'Fix': 'Use: get_all_pending_fines() instead',
                'Status': 'Need to Fix'
            },
            {
                'Test ID': 'TC-024',
                'Test Name': 'test_overdue_fine_report',
                'Module': 'test_fine_controller',
                'Error Type': 'AttributeError',
                'Error Message': "'FineController' object has no attribute 'get_unpaid_fines'",
                'Root Cause': 'Method name mismatch',
                'Fix': 'Use: get_all_fines() or get_all_pending_fines()',
                'Status': 'Need to Fix'
            }
        ]
    
    def create_excel_workbook(self):
        """Create Excel workbook with failure report"""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Test Failures'
        
        # Define header style - Red background for failures
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        
        # Define border style
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Add header row
        headers = ['Test ID', 'Test Name', 'Module', 'Error Type', 'Error Message', 'Root Cause', 'Fix', 'Status']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Add test failure data
        for row_idx, test_case in enumerate(self.failed_tests, 2):
            ws.cell(row=row_idx, column=1, value=test_case['Test ID']).border = border
            ws.cell(row=row_idx, column=2, value=test_case['Test Name']).border = border
            ws.cell(row=row_idx, column=3, value=test_case['Module']).border = border
            ws.cell(row=row_idx, column=4, value=test_case['Error Type']).border = border
            ws.cell(row=row_idx, column=5, value=test_case['Error Message']).border = border
            ws.cell(row=row_idx, column=6, value=test_case['Root Cause']).border = border
            ws.cell(row=row_idx, column=7, value=test_case['Fix']).border = border
            ws.cell(row=row_idx, column=8, value=test_case['Status']).border = border
            
            # Align all cells
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 15
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, len(self.failed_tests) + 2):
            ws.row_dimensions[row_idx].height = 60
        
        return wb
    
    def generate_failure_report(self):
        """Generate test failure report"""
        print("\n" + "="*70)
        print("Generating Test Failure Report")
        print("="*70)
        
        print(f"\nFound {len(self.failed_tests)} failed test cases")
        
        # Create workbook
        print("\nCreating test_failure_report.xlsx...")
        wb = self.create_excel_workbook()
        
        failure_report_path = os.path.join(self.output_dir, 'test_failure_report.xlsx')
        wb.save(failure_report_path)
        print(f"✓ Created: {failure_report_path}")
        
        print(f"\n✓ Test failure report generated in: {self.output_dir}")
        print("="*70 + "\n")


def main():
    """Main entry point"""
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    generator = TestFailureReportGenerator(project_root)
    generator.generate_failure_report()


if __name__ == '__main__':
    main()
