#!/usr/bin/env python
"""
Export pytest controller tests to Excel file
Generates an Excel report with test results matching the template format
"""

import subprocess
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import os


def get_test_description(test_method):
    """Get detailed test description with step-by-step breakdown"""
    descriptions = {
        # Auth Controller Tests
        'test_admin_login_valid': '1. Create admin user\n2. Input correct admin credentials\n3. Verify login success\n4. Check admin role is returned',
        'test_admin_login_invalid_password': '1. Create admin user\n2. Input admin username with wrong password\n3. Verify login fails\n4. Check error message',
        'test_admin_login_invalid_username': '1. Input non-existent admin username\n2. Input any password\n3. Verify login fails\n4. Check error message',
        'test_admin_login_empty_credentials': '1. Leave username and password empty\n2. Attempt login\n3. Verify login fails\n4. Check validation error',
        'test_librarian_login_valid': '1. Create librarian user\n2. Input correct librarian credentials\n3. Verify login success\n4. Check librarian role is returned',
        'test_librarian_login_invalid_password': '1. Create librarian user\n2. Input librarian username with wrong password\n3. Verify login fails\n4. Check error message',
        'test_librarian_login_invalid_username': '1. Input non-existent librarian username\n2. Input any password\n3. Verify login fails\n4. Check error message',
        'test_librarian_case_sensitive': '1. Create librarian user\n2. Test with different case combinations\n3. Verify case sensitivity is enforced\n4. Check only exact match works',
        'test_member_login_valid': '1. Create member user with email and password\n2. Input member email and password\n3. Verify login success\n4. Check member role and email returned',
        'test_member_login_nonexistent_email': '1. Input non-existent email\n2. Input any password\n3. Verify login fails\n4. Check "user not found" message',
        'test_member_login_wrong_password': '1. Create member user\n2. Input correct email with wrong password\n3. Verify login fails\n4. Check "invalid password" message',
        'test_member_login_deleted_account': '1. Create member user\n2. Mark user as deleted\n3. Attempt login with credentials\n4. Verify login fails with "account deleted" message',
        'test_authenticate_with_role_admin': '1. Call authenticate with admin role specified\n2. Input admin credentials\n3. Verify authentication succeeds\n4. Check admin role in response',
        'test_authenticate_with_role_librarian': '1. Call authenticate with librarian role specified\n2. Input librarian credentials\n3. Verify authentication succeeds\n4. Check librarian role in response',
        'test_authenticate_with_role_member': '1. Create member in database\n2. Call authenticate with member role\n3. Input member email and password\n4. Verify authentication succeeds and member role returned',
        'test_authenticate_admin_auto_detection': '1. Call authenticate with admin username\n2. System detects admin automatically\n3. Verify admin role is set\n4. Check authentication succeeds',
        'test_authenticate_librarian_auto_detection': '1. Call authenticate with librarian username\n2. System detects librarian automatically\n3. Verify librarian role is set\n4. Check authentication succeeds',
        'test_authenticate_invalid_role': '1. Call authenticate with invalid role parameter\n2. Input any credentials\n3. Verify authentication fails\n4. Check error for invalid role',
        'test_get_user_role_admin': '1. Call get_user_role with admin username\n2. System queries user database\n3. Verify admin role is returned\n4. Check no authentication needed',
        'test_get_user_role_librarian': '1. Call get_user_role with librarian username\n2. System queries user database\n3. Verify librarian role is returned\n4. Check no authentication needed',
        'test_change_member_password_success': '1. Create member user with password\n2. Call change_member_password with old and new password\n3. Verify old password is correct\n4. Hash and save new password\n5. Verify new password works on next login',
        
        # User Controller Tests
        'test_register_user_success': '1. Input new user details (name, email, password)\n2. Verify email not already registered\n3. Hash password using pbkdf2_hmac\n4. Save user to database\n5. Verify registration succeeds',
        'test_register_user_duplicate_email': '1. Create first user with email\n2. Attempt to register second user with same email\n3. Verify registration fails\n4. Check "duplicate email" error message',
        'test_login_valid': '1. Create user in database\n2. Input user email and password\n3. Retrieve user from database\n4. Verify password matches\n5. Check login succeeds and user data returned',
        'test_login_invalid_email': '1. Input non-existent email\n2. Input any password\n3. Query database for user\n4. Verify user not found\n5. Check login fails',
        'test_login_wrong_password': '1. Create user in database\n2. Input correct email with wrong password\n3. Retrieve user from database\n4. Verify password does not match\n5. Check login fails with error',
        'test_get_user': '1. Create user in database\n2. Call get_user with user ID\n3. Query database for user\n4. Verify user data is returned\n5. Check all fields match',
        'test_get_nonexistent_user': '1. Call get_user with invalid user ID\n2. Query database for user\n3. Verify user not found\n4. Check null is returned',
        'test_update_user': '1. Create user in database\n2. Update user details (name, email, etc)\n3. Save changes to database\n4. Retrieve updated user\n5. Verify all changes persisted',
        
        # Book Controller Tests
        'test_add_book_success': '1. Input book details (title, author, quantity)\n2. Verify title not empty\n3. Verify quantity >= 0\n4. Create book record\n5. Save to database\n6. Verify book is added',
        'test_add_book_zero_quantity': '1. Input book with zero quantity\n2. Verify quantity validation\n3. Check system handles zero quantity\n4. Save book record\n5. Verify book is created',
        'test_get_book_by_id': '1. Create book in database\n2. Call get_book with book ID\n3. Query database\n4. Verify book details match\n5. Check all fields returned',
        'test_get_nonexistent_book': '1. Call get_book with invalid ID\n2. Query database\n3. Verify book not found\n4. Check null is returned',
        'test_get_all_books': '1. Create multiple books in database\n2. Call get_all_books\n3. Query all active books\n4. Verify all books returned\n5. Check count matches',
        'test_search_books': '1. Create books with various titles\n2. Call search with search criteria\n3. Query database with filter\n4. Verify matching books returned\n5. Check non-matching books excluded',
        'test_search_books_no_results': '1. Call search with non-matching criteria\n2. Query database\n3. Verify no results found\n4. Check empty list is returned',
        'test_update_book': '1. Create book in database\n2. Update book fields (title, quantity)\n3. Save changes to database\n4. Retrieve updated book\n5. Verify all changes persisted',
        'test_delete_book': '1. Create book in database\n2. Call delete_book with book ID\n3. Mark book as deleted (soft delete)\n4. Save changes\n5. Verify book is_active = False',
        
        # Borrow Controller Tests
        'test_borrow_book_success': '1. Create user and book\n2. Call borrow_book with user and book ID\n3. Verify book available quantity > 0\n4. Create borrow record\n5. Decrease book availability\n6. Save to database',
        'test_borrow_nonexistent_book': '1. Create user\n2. Call borrow_book with non-existent book ID\n3. Query database for book\n4. Verify book not found\n5. Check borrow fails with error',
        'test_return_book_success': '1. Create borrow record\n2. Call return_book with borrow ID\n3. Mark borrow as returned\n4. Calculate return date\n5. Check for overdue and fine\n6. Update database',
        'test_return_nonexistent_borrow': '1. Call return_book with invalid borrow ID\n2. Query database\n3. Verify borrow record not found\n4. Check error is returned',
        'test_get_user_borrows': '1. Create user with multiple borrows\n2. Call get_user_borrows with user ID\n3. Query all borrow records for user\n4. Verify all borrows returned\n5. Check count matches',
        'test_get_active_borrows': '1. Create some returned and some unreturned borrows\n2. Call get_active_borrows\n3. Query unreturned borrow records\n4. Verify only active borrows returned\n5. Check returned borrows excluded',
        'test_get_overdue_borrows': '1. Create borrows with past due dates\n2. Call get_overdue_borrows\n3. Query overdue borrow records\n4. Check current date vs due date\n5. Verify overdue borrows returned',
        
        # Fine Controller Tests
        'test_get_fine': '1. Create fine record\n2. Call get_fine with fine ID\n3. Query database\n4. Verify fine details returned\n5. Check amount and status',
        'test_get_nonexistent_fine': '1. Call get_fine with invalid ID\n2. Query database\n3. Verify fine not found\n4. Check null is returned',
        'test_get_user_fines': '1. Create user with multiple fines\n2. Call get_user_fines with user ID\n3. Query all fines for user\n4. Verify all fines returned\n5. Check count matches',
        'test_get_pending_fines': '1. Create some paid and unpaid fines\n2. Call get_pending_fines\n3. Filter unpaid fines only\n4. Verify only pending returned\n5. Check paid fines excluded',
        'test_create_fine': '1. Create overdue borrow record\n2. Call create_fine\n3. Calculate fine amount (days overdue * rate)\n4. Create fine record\n5. Save to database\n6. Mark fine as unpaid',
        'test_pay_fine': '1. Create unpaid fine\n2. Call pay_fine with fine ID\n3. Record payment amount\n4. Mark fine as paid\n5. Save payment record\n6. Update database',
        'test_get_fine_statistics': '1. Create multiple fines\n2. Call get_fine_statistics\n3. Calculate total pending fines\n4. Calculate total paid\n5. Return statistics dictionary',
        
        # Report Controller Tests
        'test_get_borrow_statistics': '1. Query all borrow records\n2. Count total borrows\n3. Count active borrows\n4. Count returned borrows\n5. Calculate statistics\n6. Return report',
        'test_get_fine_statistics': '1. Query all fines\n2. Calculate total fine amount\n3. Count paid/unpaid\n4. Calculate average fine\n5. Return statistics',
        'test_get_user_statistics': '1. Query all users\n2. Count by role (admin, librarian, member)\n3. Count active/deleted\n4. Calculate statistics\n5. Return report',
        'test_get_book_statistics': '1. Query all books\n2. Count total books\n3. Calculate available copies\n4. Count by category\n5. Return statistics',
        
        # Integration Tests
        'test_complete_workflow': '1. Register new user\n2. User login\n3. Browse books\n4. Borrow book\n5. Return book\n6. View history\n7. Verify all steps complete',
        'test_auth_with_user_controller': '1. Create user via UserController\n2. Authenticate via AuthController\n3. Verify both systems work together\n4. Check user data consistency',
        
        # Error Handling Tests
        'test_user_controller_error_handling': '1. Input invalid data\n2. Attempt user operations\n3. Verify errors caught\n4. Check error messages returned',
        'test_book_controller_error_handling': '1. Input invalid book data\n2. Attempt operations\n3. Verify errors caught\n4. Check system stability',
        'test_borrow_controller_error_handling': '1. Test with invalid inputs\n2. Attempt borrow operations\n3. Verify all errors handled\n4. Check no system crash',
        'test_auth_controller_error_handling': '1. Input malformed credentials\n2. Attempt authentication\n3. Verify errors caught\n4. Check error messages clear',
    }
    
    return descriptions.get(test_method, f"Execute {test_method.replace('test_', '').replace('_', ' ').title()}")


def run_pytest():
    """Run pytest and parse output directly"""
    try:
        result = subprocess.run(
            ['python', '-m', 'pytest', 'tests/test_controller.py', '-v', '--tb=line'],
            capture_output=True,
            text=True
        )
        
        return parse_pytest_output(result.stdout)
    except Exception as e:
        print(f"Error running pytest: {e}")
        return []


def parse_pytest_output(stdout):
    """Parse pytest output to extract test results"""
    tests = []
    test_counter = {}
    
    # Parse PASSED and FAILED tests
    pattern = r'(tests/[^\s]+::[^\s]+::[^\s]+)\s+(PASSED|FAILED)'
    
    for match in re.finditer(pattern, stdout):
        test_path = match.group(1)
        status = match.group(2)
        
        # Extract test details
        parts = test_path.split('::')
        if len(parts) >= 3:
            test_class = parts[1]
            test_method = parts[2]
            
            # Generate test ID based on class
            class_prefix = test_class[:3].upper()
            if class_prefix not in test_counter:
                test_counter[class_prefix] = 1
            else:
                test_counter[class_prefix] += 1
            
            test_id = f"TC-{class_prefix}-{test_counter[class_prefix]:02d}"
            
            # Get detailed description
            test_steps = get_test_description(test_method)
            
            tests.append({
                'id': test_id,
                'name': test_method.replace('test_', '').replace('_', ' ').title(),
                'method': test_method,
                'status': status,
                'steps': test_steps,
                'details': ''
            })
    
    return tests


def create_excel_report(tests):
    """Create Excel file with test results"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Controller Tests"
    
    # Define colors
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ['Test ID', 'Test Name', 'Test Steps', 'Status', 'Details']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add test data
    for row_idx, test in enumerate(tests, 2):
        # Test ID
        cell = ws.cell(row=row_idx, column=1)
        cell.value = test['id']
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='top')
        
        # Test Name
        cell = ws.cell(row=row_idx, column=2)
        cell.value = test['name']
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Test Steps
        cell = ws.cell(row=row_idx, column=3)
        cell.value = test['steps']
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Status
        cell = ws.cell(row=row_idx, column=4)
        cell.value = test['status']
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply color based on status
        if test['status'] == 'PASSED':
            cell.fill = green_fill
        
        # Details
        cell = ws.cell(row=row_idx, column=5)
        cell.value = test['details']
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12 # type: ignore
    ws.column_dimensions['B'].width = 35 # pyright: ignore[reportOptionalMemberAccess]
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    # Set row heights for content
    for row in range(2, len(tests) + 2):
        ws.row_dimensions[row].height = 25
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 25
    
    total_tests = len(tests)
    passed_tests = sum(1 for t in tests if t['status'] == 'PASSED')
    failed_tests = sum(1 for t in tests if t['status'] == 'FAILED')
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    
    # Summary content
    summary_data = [
        ["Test Execution Summary"],
        ["Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Tests", total_tests],
        ["Passed", passed_tests],
        ["Failed", failed_tests],
        ["Pass Rate %", f"{pass_rate:.1f}%"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = yellow_fill
            cell.border = border
    
    return wb


def main():
    """Main function"""
    print("Starting test export to Excel...")
    print("Running pytest tests/test_controller.py...")
    
    # Run pytest and get results
    tests = run_pytest()
    
    if not tests:
        print("Error: Could not parse test results")
        return False
    
    print(f"Found {len(tests)} controller tests")
    
    # Create Excel file
    print("Creating Excel file...")
    wb = create_excel_report(tests)
    
    # Save file
    output_file = "Controller_Test_Results.xlsx"
    wb.save(output_file)
    
    print(f"✓ Test results exported to: {output_file}")
    
    # Print summary
    passed = sum(1 for t in tests if t['status'] == 'PASSED')
    failed = sum(1 for t in tests if t['status'] == 'FAILED')
    total = len(tests)
    pass_rate = (passed / total * 100) if total > 0 else 0
    
    print(f"\n{'='*40}")
    print(f"Test Summary")
    print(f"{'='*40}")
    print(f"Total Tests:  {total}")
    print(f"Passed:       {passed}")
    print(f"Failed:       {failed}")
    print(f"Pass Rate:    {pass_rate:.1f}%")
    print(f"{'='*40}")
    
    return True


if __name__ == "__main__":
    main()
