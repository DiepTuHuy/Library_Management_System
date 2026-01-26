#!/usr/bin/env python
"""
Export all pytest tests to Excel file
Generates comprehensive Excel reports for all tests
"""

import subprocess
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


def get_test_description(test_method):
    """Get detailed test description with step-by-step breakdown"""
    descriptions = {
        # Auth Controller Tests
        'test_admin_login_valid': '1. Create admin user\n2. Input correct admin credentials\n3. Verify login success\n4. Check admin role is returned',
        'test_member_login_valid': '1. Create member user with email and password\n2. Input member email and password\n3. Verify login success\n4. Check member role and email returned',
        'test_authenticate_with_role_member': '1. Create member in database\n2. Call authenticate with member role\n3. Input member email and password\n4. Verify authentication succeeds and member role returned',
        'test_change_member_password_success': '1. Create member user with password\n2. Call change_member_password with old and new password\n3. Verify old password is correct\n4. Hash and save new password\n5. Verify new password works on next login',
        
        # User Model Tests
        'test_create_user_basic': '1. Input user details (name, email, password)\n2. Set role to member\n3. Create user instance\n4. Verify all fields initialized\n5. Check is_active = True',
        'test_save_user_to_database': '1. Create user instance\n2. Hash password using pbkdf2_hmac\n3. Call user.save(db)\n4. Insert into database\n5. Verify user_id assigned',
        'test_update_user_in_database': '1. Create and save user\n2. Update user fields\n3. Call save again\n4. Update database record\n5. Verify changes persisted',
        'test_delete_user_soft_delete': '1. Create and save user\n2. Call user.delete()\n3. Set is_active = False\n4. Save to database\n5. Verify soft delete (not removed)',
        'test_get_user_by_id': '1. Create and save user\n2. Call User.get_by_id(db, user_id)\n3. Query database\n4. Create User object from data\n5. Verify all fields match',
        'test_get_user_by_email': '1. Create and save user with email\n2. Call User.get_by_email(db, email)\n3. Query database by email\n4. Verify only active users returned\n5. Check user data matches',
        
        # Book Model Tests
        'test_add_book_success': '1. Input book details (title, author, quantity)\n2. Verify title not empty\n3. Create book instance\n4. Call book.save(db)\n5. Insert into database\n6. Verify book_id assigned',
        'test_get_book_by_id': '1. Create and save book\n2. Call Book.get_by_id(db, book_id)\n3. Query database\n4. Create Book object\n5. Verify all fields match',
        'test_get_all_books': '1. Create multiple books\n2. Save all to database\n3. Call Book.get_all(db)\n4. Query all active books\n5. Verify count and data match',
        'test_update_book': '1. Create and save book\n2. Update book fields (title, quantity)\n3. Call save again\n4. Update database record\n5. Verify changes persisted',
        
        # Borrow Model Tests
        'test_borrow_book_success': '1. Create user and book\n2. Create borrow record\n3. Set borrow_date and due_date\n4. Call borrow.save(db)\n5. Insert into database\n6. Verify borrow_id assigned',
        'test_return_book_success': '1. Create and save borrow record\n2. Set return_date\n3. Calculate overdue days\n4. Call borrow.save()\n5. Update database\n6. Verify returned status',
        'test_get_user_borrows': '1. Create user with multiple borrows\n2. Call BorrowRecord.get_user_records(db, user_id)\n3. Query all borrow records\n4. Verify all borrows returned\n5. Check count matches',
        'test_get_active_borrows': '1. Create returned and unreturned borrows\n2. Call BorrowRecord.get_active(db, user_id)\n3. Filter where return_date is None\n4. Verify only active borrows\n5. Check returned excluded',
        'test_get_overdue_borrows': '1. Create borrows with past due dates\n2. Call BorrowRecord.get_overdue(db)\n3. Compare due_date with today\n4. Filter overdue records\n5. Verify only overdue returned',
        
        # Fine Model Tests
        'test_get_fine': '1. Create fine record\n2. Call Fine.get_by_id(db, fine_id)\n3. Query database\n4. Create Fine object\n5. Verify amount and status',
        'test_create_fine': '1. Create overdue borrow\n2. Calculate fine amount (days * rate)\n3. Create fine instance\n4. Call fine.save(db)\n5. Insert into database\n6. Mark as unpaid',
        'test_pay_fine': '1. Create unpaid fine\n2. Create payment record\n3. Set payment_date\n4. Mark fine as paid\n5. Update database\n6. Verify payment saved',
        'test_get_user_fines': '1. Create user with fines\n2. Call Fine.get_user_fines(db, user_id)\n3. Query all fines for user\n4. Verify all fines returned\n5. Check count matches',
        
        # GUI Integration Tests
        'test_login_view_initialization': '1. Create login view window\n2. Initialize form elements\n3. Set button callbacks\n4. Display username field\n5. Display password field\n6. Show login button',
        'test_login_form_validation': '1. Input empty username\n2. Check validation error\n3. Input empty password\n4. Check validation error\n5. Input valid credentials\n6. Verify validation passes',
        'test_admin_dashboard_loads': '1. Authenticate as admin\n2. Load admin dashboard\n3. Display user management section\n4. Display book management section\n5. Display reports section\n6. Verify all widgets loaded',
        'test_librarian_dashboard_loads': '1. Authenticate as librarian\n2. Load librarian dashboard\n3. Display borrow management\n4. Display return management\n5. Display fine overview\n6. Verify layout correct',
        'test_student_dashboard_loads': '1. Authenticate as member\n2. Load member dashboard\n3. Display book catalog\n4. Display my borrows\n5. Display my fines\n6. Display profile section',
        'test_navigation_between_views': '1. Load first view\n2. Click navigation button\n3. Load second view\n4. Click back button\n5. Return to first view\n6. Verify state preserved',
        'test_logout_clears_user_session': '1. Login user\n2. Store session data\n3. Click logout button\n4. Clear all session variables\n5. Return to login view\n6. Verify session empty',
        'test_books_display_in_catalog': '1. Query books from database\n2. Create book widgets\n3. Display in catalog view\n4. Verify all books shown\n5. Check pagination working\n6. Display correct details',
        'test_user_borrows_display': '1. Query user borrows\n2. Get book details for each\n3. Create borrow item widgets\n4. Display in list\n5. Show due dates\n6. Verify count correct',
        'test_fines_display_in_dashboard': '1. Query user fines\n2. Get fine details\n3. Create fine widgets\n4. Display pending vs paid\n5. Show amounts\n6. Verify layout clear',
        
        # Workflow Tests
        'test_user_registration_and_login': '1. Register new user\n2. Input email and password\n3. Hash and save to database\n4. Login with credentials\n5. Verify user retrieved\n6. Check session created',
        'test_full_borrow_return_cycle': '1. Browse catalog\n2. Select book\n3. Borrow book\n4. Decrease availability\n5. Return book\n6. Check fine calculation\n7. Update availability',
        'test_overdue_book_generates_fine': '1. Create borrow record\n2. Set past due_date\n3. Check for overdue\n4. Calculate fine amount\n5. Create fine record\n6. Verify fine in database\n7. Check amount calculated',
        'test_build_library_catalog': '1. Create multiple books\n2. Add to database\n3. Organize by category\n4. Build search index\n5. Load into catalog\n6. Verify all books display\n7. Check filtering works',
    }
    
    return descriptions.get(test_method, f"Execute {test_method.replace('test_', '').replace('_', ' ').title()}")


def run_pytest_all():
    """Run all pytest tests"""
    try:
        result = subprocess.run(
            ['python', '-m', 'pytest', 'tests/', '-v', '--tb=line'],
            capture_output=True,
            text=True
        )
        
        return parse_pytest_output(result.stdout)
    except Exception as e:
        print(f"Error running pytest: {e}")
        return []


def parse_pytest_output(stdout):
    """Parse pytest output to extract test results"""
    tests = []
    test_counter = {}
    
    # Parse PASSED and FAILED tests
    pattern = r'(tests/[^\s]+::[^\s]+::[^\s]+)\s+(PASSED|FAILED)'
    
    for match in re.finditer(pattern, stdout):
        test_path = match.group(1)
        status = match.group(2)
        
        # Extract test details
        parts = test_path.split('::')
        if len(parts) >= 3:
            file_path = parts[0]
            test_class = parts[1]
            test_method = parts[2]
            
            # Get file type
            if 'integration' in file_path:
                file_type = 'INT'
            elif 'unit' in file_path:
                file_type = 'UNT'
            else:
                file_type = 'CTL'
            
            # Generate test ID
            class_prefix = test_class[:2].upper()
            counter_key = f"{file_type}_{class_prefix}"
            
            if counter_key not in test_counter:
                test_counter[counter_key] = 1
            else:
                test_counter[counter_key] += 1
            
            test_id = f"{file_type}-{class_prefix}-{test_counter[counter_key]:03d}"
            
            # Get detailed description
            test_steps = get_test_description(test_method)
            
            tests.append({
                'id': test_id,
                'name': test_method.replace('test_', '').replace('_', ' ').title(),
                'method': test_method,
                'file_type': file_type,
                'file_path': file_path,
                'status': status,
                'steps': test_steps,
                'details': ''
            })
    
    return tests


def create_excel_report(tests):
    """Create Excel file with test results"""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Tests"
    
    # Define colors
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ['Test ID', 'Test Name', 'Test Steps', 'Status', 'Details']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = yellow_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add test data
    for row_idx, test in enumerate(tests, 2):
        # Test ID
        cell = ws.cell(row=row_idx, column=1)
        cell.value = test['id']
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='top')
        
        # Test Name
        cell = ws.cell(row=row_idx, column=2)
        cell.value = test['name']
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Test Steps
        cell = ws.cell(row=row_idx, column=3)
        cell.value = test['steps']
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Status
        cell = ws.cell(row=row_idx, column=4)
        cell.value = test['status']
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply color based on status
        if test['status'] == 'PASSED':
            cell.fill = green_fill
        
        # Details
        cell = ws.cell(row=row_idx, column=5)
        cell.value = test['details']
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    # Set row heights for content
    for row in range(2, len(tests) + 2):
        ws.row_dimensions[row].height = 20
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 25
    
    total_tests = len(tests)
    passed_tests = sum(1 for t in tests if t['status'] == 'PASSED')
    failed_tests = sum(1 for t in tests if t['status'] == 'FAILED')
    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    
    # Count by type
    unit_tests = sum(1 for t in tests if t['file_type'] == 'UNT')
    integration_tests = sum(1 for t in tests if t['file_type'] == 'INT')
    controller_tests = sum(1 for t in tests if t['file_type'] == 'CTL')
    
    # Summary content
    summary_data = [
        ["Test Execution Summary"],
        ["Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        [""],
        ["Overall Results"],
        ["Total Tests", total_tests],
        ["Passed", passed_tests],
        ["Failed", failed_tests],
        ["Pass Rate %", f"{pass_rate:.1f}%"],
        [""],
        ["Tests by Type"],
        ["Unit Tests", unit_tests],
        ["Integration Tests", integration_tests],
        ["Controller Tests", controller_tests],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            if row_idx in [1, 4, 10]:
                cell.font = Font(bold=True, size=12)
                cell.fill = yellow_fill
            cell.border = border
    
    return wb


def main():
    """Main function"""
    print("\n" + "="*50)
    print("  TEST EXPORT TO EXCEL - ALL TESTS")
    print("="*50 + "\n")
    
    print("Running pytest all tests...")
    
    # Run pytest and get results
    tests = run_pytest_all()
    
    if not tests:
        print("Error: Could not parse test results")
        return False
    
    print(f"Found {len(tests)} total tests\n")
    
    # Create Excel file
    print("Creating Excel file...")
    wb = create_excel_report(tests)
    
    # Save file
    output_file = "All_Test_Results.xlsx"
    wb.save(output_file)
    
    print(f"✓ Test results exported to: {output_file}\n")
    
    # Print summary
    passed = sum(1 for t in tests if t['status'] == 'PASSED')
    failed = sum(1 for t in tests if t['status'] == 'FAILED')
    unit = sum(1 for t in tests if t['file_type'] == 'UNT')
    integration = sum(1 for t in tests if t['file_type'] == 'INT')
    controller = sum(1 for t in tests if t['file_type'] == 'CTL')
    total = len(tests)
    pass_rate = (passed / total * 100) if total > 0 else 0
    
    print("="*50)
    print("Test Summary")
    print("="*50)
    print(f"Total Tests:        {total}")
    print(f"Passed:             {passed}")
    print(f"Failed:             {failed}")
    print(f"Pass Rate:          {pass_rate:.1f}%")
    print("\nTests by Type:")
    print(f"  Unit Tests:       {unit}")
    print(f"  Integration:      {integration}")
    print(f"  Controller:       {controller}")
    print("="*50 + "\n")
    
    return True


if __name__ == "__main__":
    main()
